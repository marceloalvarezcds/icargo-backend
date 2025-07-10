import os
from datetime import datetime
from decimal import Decimal
from typing import List, Optional

from fastapi import HTTPException  # type: ignore
from jinja2 import Template
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from pdfkit import from_string  # type: ignore
from sqlalchemy.orm import Session  # type: ignore
from fastapi.responses import HTMLResponse

from app import repositories, schemas
from app.config import LOGO_IMAGE_URL, REPORTS_FOLDER, templateEnv
from app.enums import (
    LiquidacionEstadoEnum,
    LiquidacionEtapaEnum,
    MovimientoEstadoEnum,
    TipoContraparteEnum,
    EstadoEnum
)
from app.models import (
    Camion,
    Instrumento,
    Liquidacion,
    Movimiento,
    OrdenCargaAnticipoRetirado,
)
from app.schemas import (
    LiquidacionAddInstrumentosForm,
    LiquidacionAddMovimientosForm,
    LiquidacionForm,
    LiquidacionCabeceraMovimientosForm,
    InstrumentoForm,
    Instrumento,
    LiquidacionReport
)
from app.schemas import Instrumento as InstrumentoSchema
from app.utils import get_gestor_carga_by_params, number_format

from .camion import update_camion_anticipo_retirado
from .instrumento import create_instrumento
from .user import get_user_by_username
from app.logger import logger


def get_liquidacion_list(
    db: Session, gestor_carga_id: Optional[int]
) -> List[Liquidacion]:
    if gestor_carga_id:
        return repositories.get_liquidacion_list_by_gestor_carga_id(db, gestor_carga_id)
    return repositories.get_liquidacion_list(db)


def get_liquidacion_list_by_estado_cuenta(
    db: Session,
    tipo_contraparte_id: int,
    contraparte_id: int,
    contraparte: str,
    contraparte_numero_documento: str,
    etapa: str,
    gestor_carga_id: Optional[int],
) -> List[Liquidacion]:
    if gestor_carga_id:
        return repositories.get_liquidacion_list_by_contraparte_and_gestor_carga_id(
            db,
            tipo_contraparte_id,
            contraparte_id,
            contraparte,
            contraparte_numero_documento,
            etapa,
            gestor_carga_id,
        )
    return repositories.get_liquidacion_list_by_contraparte(
        db,
        tipo_contraparte_id,
        contraparte_id,
        contraparte,
        contraparte_numero_documento,
        etapa,
    )


def create_liquidacion(
    db: Session,
    data: LiquidacionForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Liquidacion:
    gestor_id = get_gestor_carga_by_params(data, gestor_carga_id)
    return repositories.create_liquidacion(db, data, gestor_id, modified_by)


def create_liquidacion_pendiente(
    db: Session,
    data: LiquidacionCabeceraMovimientosForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Liquidacion:
    gestor_id = None
    chofer_id = None
    propietario_id = None
    proveedor_id = None
    remitente_id = None
    punto_venta = None
    nombre_contraparte = None
    contraparte_documento = None
    moneda = None
# camino con movimientos
    if len(data.movimientos) > 0:
        movimientos = get_movimiento_list_by_liquidacion_create_form(db, data)
        movimiento = movimientos[0]
        gestor_id = get_gestor_carga_by_params(movimiento, gestor_carga_id)
        nombre_contraparte = movimiento.contraparte
        contraparte_documento = movimiento.contraparte_numero_documento
        moneda = data.moneda.id
        if movimiento.es_chofer:
            chofer_id = movimiento.chofer_id
        elif movimiento.es_proveedor:
            proveedor_id = movimiento.proveedor_id
            if data.cabecera.punto_venta_id:
                punto_venta = data.cabecera.punto_venta_id
                nombre_contraparte = data.cabecera.contraparte_pdv
                contraparte_documento = data.cabecera.contraparte_numero_documento_pdv
        elif movimiento.es_gestor:
            remitente_id = movimiento.remitente_id
        else:
            propietario_id = movimiento.propietario_id
    else:
        moneda = data.moneda.id
        nombre_contraparte = data.cabecera.contraparte
        contraparte_documento = data.cabecera.contraparte_numero_documento
        if data.cabecera.tipo_contraparte_descripcion == TipoContraparteEnum.CHOFER.value:
            chofer_id = data.cabecera.contraparte_id
        elif data.cabecera.tipo_contraparte_descripcion == TipoContraparteEnum.REMITENTE.value:
            remitente_id = data.cabecera.contraparte_id
        elif data.cabecera.tipo_contraparte_descripcion == TipoContraparteEnum.PROVEEDOR.value:
            proveedor_id = data.cabecera.contraparte_id
            if data.cabecera.punto_venta_id:
                punto_venta = data.cabecera.punto_venta_id
                nombre_contraparte = data.cabecera.contraparte_pdv
                contraparte_documento = data.cabecera.contraparte_numero_documento_pdv
        elif data.cabecera.tipo_contraparte_descripcion == TipoContraparteEnum.PUNTO_VENTA.value:
            proveedor_id = data.cabecera.contraparte_id
            punto_venta = data.cabecera.punto_venta_id
            nombre_contraparte = data.cabecera.contraparte_pdv
            contraparte_documento = data.cabecera.contraparte_numero_documento_pdv
        elif data.cabecera.tipo_contraparte_descripcion == TipoContraparteEnum.PROPIETARIO.value:
            propietario_id = data.cabecera.contraparte_id

    # logger.info("create_liquidacion_pendiente")
    # logger.info(f"tipo_contraparte_descripcion {movimiento.tipo_contraparte_descripcion}")

    liquidacion = create_liquidacion(
        db,
        LiquidacionForm(
            tipo_contraparte_id=data.cabecera.tipo_contraparte_id,
            contraparte=nombre_contraparte,
            contraparte_numero_documento=contraparte_documento,
            moneda_id= moneda,
            # IDs para referencia a las tablas de las contraparte
            chofer_id=chofer_id,
            gestor_carga_id=gestor_carga_id,
            propietario_id=propietario_id,
            proveedor_id=proveedor_id,
            remitente_id=remitente_id,
            punto_venta_id=punto_venta,
            es_pago_cobro=data.es_pago_cobro,
            monto=data.monto,
            tipo_mov_liquidacion=data.tipo_mov_liquidacion,
            es_orden_pago=data.es_orden_pago
        ),
        gestor_id,
        modified_by,
    )
    if len(data.movimientos) > 0:
        liquidacion.movimientos = movimientos
    db.commit()
    return liquidacion


def get_liquidacion_by_id(db: Session, id: int) -> Liquidacion:
    obj = repositories.get_liquidacion_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Liquidación no encontrada")
    return obj


def edit_liquidacion(
    id: int,
    db: Session,
    data: LiquidacionForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Liquidacion:
    gestor_id = get_gestor_carga_by_params(data, gestor_carga_id)
    to_edit_obj = get_liquidacion_by_id(db, id)
    return repositories.edit_liquidacion(to_edit_obj, db, data, gestor_id, modified_by)


def delete_liquidacion(db: Session, id: int, modified_by: str) -> Liquidacion:
    obj = get_liquidacion_by_id(db, id)
    movimientos = repositories.get_movimiento_list_by_liquidacion_id(db, id)
    change_movimiento_list_status(
        db, movimientos, LiquidacionEstadoEnum.PENDIENTE, modified_by
    )
    obj.movimientos = []
    return repositories.delete_liquidacion(obj, db, modified_by)


def add_movimientos(
    id: int,
    db: Session,
    data: LiquidacionAddMovimientosForm,
    modified_by: str,
) -> Liquidacion:
    to_edit_obj = get_liquidacion_by_id(db, id)
    new_movimientos = get_movimiento_list_by_liquidacion_create_form(db, data)
    change_movimiento_list_status(
        db, new_movimientos, LiquidacionEstadoEnum.EN_PROCESO, modified_by
    )
    old_movimientos = to_edit_obj.movimientos
    to_edit_obj.movimientos = [*old_movimientos, *new_movimientos]
    to_edit_obj.modified_by = modified_by
    to_edit_obj.modified_at = datetime.now()
    to_edit_obj.pago_cobro =  sum(
            x.saldo_ml for x in to_edit_obj.movimientos if x.estado != EstadoEnum.ELIMINADO.value
        )
    to_edit_obj.es_pago_cobro = "PAGO" if to_edit_obj.pago_cobro > 0 else "COBRO"
    db.commit()
    return to_edit_obj


def remove_movimiento(
    id: int,
    db: Session,
    data: schemas.Movimiento,
    modified_by: str,
) -> Liquidacion:
    to_edit_obj = get_liquidacion_by_id(db, id)
    movimiento_to_remove = get_movimiento_by_schema(db, data)
    movimiento_to_remove.liquidacion_id = None
    movimiento_to_remove.estado = LiquidacionEstadoEnum.PENDIENTE.value
    movimiento_to_remove.modified_by = modified_by
    movimiento_to_remove.modified_at = datetime.now()
    db.commit()
    old_movimientos: List[Movimiento] = to_edit_obj.movimientos
    new_movimientos = list(filter(lambda x: x.id != data.id, old_movimientos))
    to_edit_obj.movimientos = new_movimientos
    to_edit_obj.modified_by = modified_by
    to_edit_obj.modified_at = datetime.now()
    to_edit_obj.pago_cobro =  sum(
            x.saldo_ml for x in to_edit_obj.movimientos if x.estado != EstadoEnum.ELIMINADO.value
        )
    to_edit_obj.es_pago_cobro = "PAGO" if to_edit_obj.pago_cobro > 0 else "COBRO"
    db.commit()
    return to_edit_obj


def refresh_pago_cobro(
    db: Session,
    id: int,
    modified_by: str,
) -> Liquidacion:
    to_edit_obj = get_liquidacion_by_id(db, id)
    movimientos = to_edit_obj.movimientos
    to_edit_obj.pago_cobro =  sum(
        x.saldo for x in movimientos if x.estado != EstadoEnum.ELIMINADO.value
    )
    db.commit()
    return to_edit_obj


def remove_movimientos(
    id: int,
    db: Session,
    data: schemas.LiquidacionAddMovimientosForm,
    modified_by: str,
) -> Liquidacion:
    to_edit_obj = get_liquidacion_by_id(db, id)
    delete_movimientos = []
    for m in data.movimientos:
        mov = get_movimiento_by_schema(db, m)
        if mov:
            mov.liquidacion_id = None
            mov.estado = LiquidacionEstadoEnum.PENDIENTE.value
            mov.modified_by = modified_by
            mov.modified_at = datetime.now()
            delete_movimientos.append(mov)

    old_movimientos: List[Movimiento] = to_edit_obj.movimientos
    result = [mov for mov in old_movimientos if mov not in delete_movimientos]
    to_edit_obj.movimientos = result
    to_edit_obj.modified_by = modified_by
    to_edit_obj.modified_at = datetime.now()
    to_edit_obj.pago_cobro =  sum(
            x.saldo_ml for x in to_edit_obj.movimientos if x.estado != EstadoEnum.ELIMINADO.value
        )
    to_edit_obj.es_pago_cobro = "PAGO" if to_edit_obj.pago_cobro > 0 else "COBRO"
    db.commit()

    return to_edit_obj


def aceptar_liquidacion(db: Session, id: int, modified_by: str) -> Liquidacion:
    obj = get_liquidacion_by_id(db, id)
    movimientos = repositories.get_movimiento_list_by_liquidacion(
        db, id, MovimientoEstadoEnum.EN_PROCESO.value
    )
    change_movimiento_list_status(
        db, movimientos, LiquidacionEstadoEnum.CONFIRMADO, modified_by
    )
    obj.etapa = LiquidacionEtapaEnum.CONFIRMADO.value
    return repositories.change_liquidacion_status(
        obj, db, LiquidacionEstadoEnum.SALDO_ABIERTO, modified_by
    )


def cancelar_liquidacion(db: Session, id: int, modified_by: str) -> Liquidacion:
    obj = get_liquidacion_by_id(db, id)
    movimientos = repositories.get_movimiento_list_by_liquidacion_id(
        db, id
    )
    change_movimiento_list_status(
        db, movimientos, LiquidacionEstadoEnum.PENDIENTE, modified_by
    )
    obj.movimientos = []
    if not obj.comentarios:
        obj.comentarios = ""
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    full_name = f"{modified_by}"
    obj.comentarios += (
        f"<strong>{full_name} ({date}): </strong><ul>Liquidacion Cancelada </ul>"
    )
    return repositories.change_liquidacion_status(
        obj, db, LiquidacionEstadoEnum.CANCELADO, modified_by
    )


def rechazar_liquidacion(
    db: Session, id: int, comentario: str, current_user: schemas.AuthUser
) -> Liquidacion:
    return change_liquidacion_status(
        db, id, comentario, LiquidacionEstadoEnum.RECHAZADO, current_user
    )


def en_revision_liquidacion(
    db: Session, id: int, comentario: str, current_user: schemas.AuthUser
) -> Liquidacion:
    return change_liquidacion_status(
        db, id, comentario, LiquidacionEstadoEnum.EN_REVISION, current_user
    )


def add_instrumento(
    id: int,
    db: Session,
    data: InstrumentoForm,
    modified_by: str,
) -> InstrumentoSchema:

    obj = get_liquidacion_by_id(db, id)
    saldo_residual = obj.saldo_residual - data.monto_ml

    if int(saldo_residual) < 0:
        raise HTTPException(
            status_code=409,
            detail="La suma de los instrumentos ha superado el saldo de la liquidacion",
        )

    instrumento = create_instrumento(db, id, data, modified_by)

    return instrumento


def add_instrumentos(
    id: int,
    db: Session,
    data: LiquidacionAddInstrumentosForm,
    modified_by: str,
) -> Liquidacion:
    saldo_residual = 0  # TODO: obtener de lo que viene del front

    if len(data.instrumentos) == 0:
        raise HTTPException(
            status_code=409, detail="Debe elegir al menos un instrumento"
        )

    if int(saldo_residual) == 0:
        obj = get_liquidacion_by_id(db, id)
        obj = repositories.change_liquidacion_status(
            obj, db, LiquidacionEstadoEnum.SALDO_CERRADO, modified_by
        )
        get_instrumento_list_by_liquidacion_create_form(db, id, data, modified_by)
        saldo_cc = data.instrumentos[0].saldo_cc
        obj.modified_by = modified_by
        obj.modified_at = datetime.now()
        obj.saldo_cc = saldo_cc
        # actualizar saldo cc liquidacion
        db.commit()
        db.refresh(obj)
        #obj = repositories.change_liquidacion_status(
        #    obj, db, LiquidacionEstadoEnum.SALDO_CERRADO, modified_by
        #)
    else:
        raise HTTPException(
            status_code=404,
            detail="La suma de los instrumentos debe ser igual al Valor de la operación",
        )
    return obj


def get_liquidacion_resumen_pdf_by_id(db: Session, id: int, estado: str) -> str:
    obj = repositories.get_liquidacion_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Orden de Carga no encontrada")
    gestor_carga = repositories.get_gestor_carga_by_id(db, obj.gestor_carga_id)
    if not gestor_carga:
        raise HTTPException(status_code=404, detail="Gestor no encontrado")
    # Obtención del usuario
    usuario = get_user_by_username(db, obj.created_by)
    usuario_nombre = f"{usuario.first_name} {usuario.last_name}" if usuario else "Sistema"
    OUTPUT_FILENAME = f"resumen_liquidacion_{id}.pdf"
    TEMPLATE_FILENAME = "pdf_liquidacion_resumen.html"
    templateEnv.filters["number_format"] = number_format
    template: Template = templateEnv.get_template(TEMPLATE_FILENAME)
    flete_movimientos: List[
        Movimiento
    ] = repositories.get_movimiento_list_for_flete_pdf_reports_by_liquidacion_id(
        db, id, estado
    )
    otro_movimientos: List[
        Movimiento
    ] = repositories.get_movimiento_list_for_otro_pdf_reports_by_liquidacion_id(
        db, id, estado
    )
    instrumentos: List[
        Instrumento
    ] = repositories.get_instrumento_list_by_liquidacion_id(db, id)
    # Obtención de totales
    # moneda = gestor_carga.moneda_simbolo
    total_contraparte = Decimal(0)
    total_flete = Decimal(0)
    total_anticipo_efectivo = Decimal(0)
    total_anticipo_combustible = Decimal(0)
    total_anticipo_otro = Decimal(0)
    total_otros = Decimal(0)

# totalizar por OC
# agregar a LiquidacionReport
# LiquidacionReport va ser una lista

    liquidacionList: List[LiquidacionReport] = []
    oc_liquidacion=LiquidacionReport(
        orden_carga_id=0,  total_orden_carga=0, movimientos=[]
    )
    oc_id: int = 0

    logger.info('******************************************************************: ')
    logger.info('******************************************************************: ')
    logger.info(f'flete_movimientos {len(flete_movimientos)}')

    for mov in flete_movimientos:

        logger.info('flete_movimientos:')
        logger.info(f'oc: {oc_id}')
        logger.info(f'mov.orden_carga_id: {mov.orden_carga_id}')

        if oc_id == 0:
            oc_liquidacion = LiquidacionReport(
                orden_carga_id=mov.orden_carga_id, total_orden_carga=0, movimientos=[]
            )
            oc_id= mov.orden_carga_id

        if oc_id != mov.orden_carga_id:
            #oc_liquidacion.total_orden_carga += total_contraparte
            liquidacionList.append(oc_liquidacion)
            oc_liquidacion = LiquidacionReport(
                orden_carga_id=mov.orden_carga_id, total_orden_carga=0, movimientos=[]
            )
            oc_id= mov.orden_carga_id

        oc_liquidacion.movimientos.append(mov)
        total_contraparte += mov.monto_mon_local
        oc_liquidacion.total_orden_carga += mov.monto_mon_local
        total_flete += (
            mov.monto_mon_local
            if (mov.es_flete or mov.es_complemento or mov.es_descuento or mov.es_merma)
            else Decimal(0)
        )
        total_anticipo_efectivo += mov.monto_mon_local if mov.es_anticipo_efectivo else Decimal(0)
        total_anticipo_combustible += (
            mov.monto_mon_local if mov.es_anticipo_combustible else Decimal(0)
        )
        total_anticipo_otro += mov.monto_mon_local if mov.es_anticipo_otro else Decimal(0)

    #oc_liquidacion.total_orden_carga += total_contraparte
    liquidacionList.append(oc_liquidacion)

    for mov in otro_movimientos:
        total_otros += mov.monto_mon_local

    # FIN Obtención de totales
    data = {
        "id": id,
        "gestor_carga_direccion": gestor_carga.direccion,
        "gestor_carga_logo": gestor_carga.logo,
        "gestor_carga_nombre": gestor_carga.nombre,
        "gestor_carga_numero_documento": gestor_carga.numero_documento,
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "usuario_nombre": usuario_nombre,
        "contraparte": obj.contraparte,
        "documento_contraparte": obj.contraparte_numero_documento,
        "tipo_contraparte": obj.tipo_contraparte_descripcion,
        "flete_movimientos": liquidacionList,
        "otro_movimientos": otro_movimientos,
        "instrumentos": instrumentos,
        "total_contraparte": f"{number_format(total_contraparte)}",
        "total_flete": f"{number_format(total_flete)}",
        "total_anticipo_efectivo": f"{number_format(total_anticipo_efectivo)}",
        "total_anticipo_combustible": f"{number_format(total_anticipo_combustible)}",
        "total_anticipo_otro": f"{number_format(total_anticipo_otro)}",
        "total_otros": f"{number_format(total_otros)}",
        "total_instrumentos": f"{number_format(obj.instrumentos_saldo)}",
    }
    #source_html = template.render(logo=LOGO_IMAGE_URL, **data)
    source_html = template.render(logo=LOGO_IMAGE_URL, times=range(2), **data)
    #logger.info(f'html: {source_html}')

    for o in liquidacionList:
        logger.info(f'orden_carga_id: {o.orden_carga_id}')
        logger.info(f'movimientos: {len(o.movimientos)}')

    pdf_filename = os.path.join(REPORTS_FOLDER, OUTPUT_FILENAME)
    from_string(source_html, pdf_filename, {"page-size": "Legal"})

    logger.info('******************************************************************: ')
    logger.info('******************************************************************: ')

    return OUTPUT_FILENAME
    #return HTMLResponse(content=source_html, status_code=200)


def get_reports(datalist: List[Liquidacion]) -> str:
    wb = Workbook()
    ws = wb.active

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "ID"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Fecha Aprobación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "Aprobador por"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Tipo de Contraparte"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Nombre Contraparte"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=6)
    title_cell.value = "Nº Contraparte"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Cobro/Pago"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=8)
    title_cell.value = "Moneda"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=9)
    title_cell.value = "Valor de operación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=10)
    title_cell.value = "Valor de instrumento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=11)
    title_cell.value = "Residuo"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=12)
    title_cell.value = "Estado"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=13)
    title_cell.value = "Fecha de Pago/Cobro"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=14)
    title_cell.value = "Fecha modificación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=15)
    title_cell.value = "Usuario modificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.id

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.created_at

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.created_by

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.tipo_contraparte_descripcion

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = item.contraparte

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.contraparte_numero_documento

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = item.tipo_operacion_descripcion

        value_cell = ws.cell(row=row + 2, column=8)
        value_cell.value = item.moneda_nombre

        value_cell = ws.cell(row=row + 2, column=9)
        value_cell.value = item.movimientos_saldo

        value_cell = ws.cell(row=row + 2, column=10)
        value_cell.value = item.instrumentos_saldo

        value_cell = ws.cell(row=row + 2, column=11)
        value_cell.value = item.saldo_residual

        value_cell = ws.cell(row=row + 2, column=12)
        value_cell.value = item.estado

        value_cell = ws.cell(row=row + 2, column=13)
        value_cell.value = item.fecha_pago_cobro

        value_cell = ws.cell(row=row + 2, column=14)
        value_cell.value = item.modified_by

        value_cell = ws.cell(row=row + 2, column=15)
        value_cell.value = item.modified_at

    ws.auto_filter.ref = ws.dimensions
    filename = "liquidacion_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename


def change_liquidacion_status(
    db: Session,
    id: int,
    comentario: str,
    estado: LiquidacionEstadoEnum,
    current_user: schemas.AuthUser,
) -> Liquidacion:
    obj = get_liquidacion_by_id(db, id)
    if comentario:
        if not obj.comentarios:
            obj.comentarios = ""
        date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        comentarios = "".join([f"<li>{c}</li>" for c in comentario.split(".")])
        full_name = f"{current_user.first_name} {current_user.last_name}"
        obj.comentarios += (
            f"<strong>{full_name} ({date}): </strong><ul>{comentarios}</ul>"
        )
    return repositories.change_liquidacion_status(
        obj, db, estado, current_user.username
    )


def change_movimiento_list_status(
    db: Session,
    movimientos: List[Movimiento],
    estado: LiquidacionEstadoEnum,
    modified_by: str,
):
    for mov in movimientos:
        mov.estado = estado.value
        mov.modified_by = modified_by
        mov.modified_at = datetime.now()
        if (
            mov.anticipo and (
                estado == LiquidacionEstadoEnum.CONFIRMADO
                or estado == LiquidacionEstadoEnum.FINALIZADO
                or estado == LiquidacionEstadoEnum.PENDIENTE
            )
        ):
            db.commit()
            anticipo: OrdenCargaAnticipoRetirado = mov.anticipo
            camion: Camion = anticipo.orden_carga.camion
            update_camion_anticipo_retirado(db, camion)
    db.commit()


def get_liquidacion_reports(db: Session) -> str:
    datalist = repositories.get_liquidacion_list(db)
    return get_reports(datalist)


def get_liquidacion_reports_by_estado_cuenta(
    db: Session,
    tipo_contraparte_id: int,
    contraparte_id: int,
    contraparte: str,
    contraparte_numero_documento: str,
    etapa: str,
    gestor_carga_id: Optional[int],
) -> str:
    datalist = get_liquidacion_list_by_estado_cuenta(
        db,
        tipo_contraparte_id,
        contraparte_id,
        contraparte,
        contraparte_numero_documento,
        etapa,
        gestor_carga_id,
    )
    return get_reports(datalist)


def get_movimiento_by_schema(db: Session, movimiento: schemas.Movimiento) -> Movimiento:
    obj = repositories.get_movimiento_by_id(db, movimiento.id)
    if not obj:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    return obj


def get_movimiento_list_by_liquidacion_create_form(
    db: Session, data: LiquidacionAddMovimientosForm
) -> List[Movimiento]:
    mList = data.movimientos
    if len(mList) == 0:
        raise HTTPException(
            status_code=409, detail="Debe elegir al menos un movimiento"
        )
    movimientos: List[Movimiento] = []
    for m in mList:
        mov = get_movimiento_by_schema(db, m)
        if mov:
            mov.estado = LiquidacionEstadoEnum.EN_PROCESO.value
            movimientos.append(mov)
    return movimientos


def get_instrumento_list_by_liquidacion_create_form(
    db: Session, id: int, data: LiquidacionAddInstrumentosForm, modified_by: str
) -> List[Instrumento]:
    iList = data.instrumentos
    if len(iList) == 0:
        raise HTTPException(
            status_code=409, detail="Debe elegir al menos un instrumento"
        )
    instrumentos: List[Instrumento] = []
    for i in iList:
        instrumento = create_instrumento(db, id, i, modified_by)
        instrumentos.append(instrumento)
    return instrumentos


def someter_liquidacion(
    db: Session, id: int, data: schemas.LiquidacionSometer, current_user: schemas.AuthUser
) -> Liquidacion:
    to_edit_obj = get_liquidacion_by_id(db, id)
    to_edit_obj.modified_by = current_user.username
    to_edit_obj.modified_at = datetime.now()
    to_edit_obj.pago_cobro = data.monto
    db.commit()
    db.refresh(to_edit_obj)
    return change_liquidacion_status(
        db, id, data.comentario, LiquidacionEstadoEnum.PENDIENTE, current_user
    )


def forzar_cierre(
    db: Session, id: int, data: schemas.LiquidacionSometer, current_user: schemas.AuthUser
) -> Liquidacion:
    to_edit_obj = get_liquidacion_by_id(db, id)

    if ( not (
            to_edit_obj.estado == LiquidacionEstadoEnum.ACEPTADO.value
            or to_edit_obj.estado == LiquidacionEstadoEnum.SALDO_ABIERTO.value
            or to_edit_obj.estado == LiquidacionEstadoEnum.SALDO_CERRADO.value
            ) ):
        raise HTTPException(
            status_code=409, detail=f"No se puede forzar cierre en estado {to_edit_obj.estado}"
        )

    to_edit_obj.modified_by = current_user.username
    to_edit_obj.modified_at = datetime.now()
    to_edit_obj.estado = LiquidacionEstadoEnum.FINALIZADO.value
    to_edit_obj.etapa = LiquidacionEstadoEnum.FINALIZADO.value

    if data.comentario:
        if not to_edit_obj.comentarios:
            to_edit_obj.comentarios = ""
        date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        comentarios = "".join([f"<li>{c}</li>" for c in data.comentario.split(".")])
        full_name = f"{current_user.first_name} {current_user.last_name}"
        to_edit_obj.comentarios += (
            f"<strong>{full_name} ({date}): </strong><ul>{comentarios}</ul>"
        )

    change_movimiento_list_status(
        db, to_edit_obj.movimientos, LiquidacionEtapaEnum.FINALIZADO, current_user.username
    )

    db.commit()
    db.refresh(to_edit_obj)
    return to_edit_obj
