from operator import and_
import os
from datetime import datetime
from fastapi import HTTPException

from typing import List, Optional
from sqlalchemy import desc, func
from sqlalchemy.orm import aliased

from sqlalchemy.orm import Session  # type: ignore

from app import repositories
from app import schemas
from app import models
from app.enums.estado import EstadoEnum
from app.models import InsumoPuntoVentaPrecio, InsumoPuntoVenta
from app.repositories import (
    crear_precio_insumo_punto_venta,
    get_insumo_punto_venta_by_insumo_id_and_moneda_id_and_punto_venta_id,
    get_last_insumo_punto_venta_precio_by_insumo_punto_venta_id,
)
from app.schemas import InsumoPuntoVentaPrecioForm
from app.utils.gestor_carga import get_gestor_carga_by_params
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from app.config import REPORTS_FOLDER


def get_insumo_venta_precio_by_id(db: Session, id: int) -> InsumoPuntoVentaPrecio:
    obj = repositories.get_insumo_venta_precio_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Insumo Venta Precio no encontrado")
    return obj


def get_insumo_punto_venta_precio_list(db: Session, gestor_carga_id: int) -> List[models.InsumoPuntoVentaPrecio]:
    return (
        db.query(models.InsumoPuntoVentaPrecio)
        .join(models.InsumoPuntoVenta)  # Unir con InsumoPuntoVenta
        .filter(
            models.InsumoPuntoVenta.gestor_carga_id == gestor_carga_id,
            models.InsumoPuntoVentaPrecio.estado == EstadoEnum.ACTIVO.value
        )
        .order_by(desc(models.InsumoPuntoVentaPrecio.id))  # Ordenar por id en orden descendente
        .all()
    )


def get_inactive_insumo_punto_venta_precio_list(db: Session, gestor_carga_id: int) -> List[models.InsumoPuntoVentaPrecio]:
    return (
        db.query(models.InsumoPuntoVentaPrecio)
        .join(models.InsumoPuntoVenta)  # Unir con InsumoPuntoVenta
        .filter(
            models.InsumoPuntoVenta.gestor_carga_id == gestor_carga_id,
            models.InsumoPuntoVentaPrecio.estado == EstadoEnum.INACTIVO.value,  # Filtrar por estado inactivo
        )
        .order_by(desc(models.InsumoPuntoVentaPrecio.id))  # Ordenar por id en orden descendente
        .all()
    )


def get_active_insumo_punto_venta_precio_list(db: Session, gestor_carga_id: int) -> List[models.InsumoPuntoVentaPrecio]:
    results = (
        db.query(models.InsumoPuntoVentaPrecio)
        .join(models.InsumoPuntoVenta)
        .filter(
            models.InsumoPuntoVenta.gestor_carga_id == gestor_carga_id,
            models.InsumoPuntoVentaPrecio.estado == EstadoEnum.ACTIVO.value
        )
        .order_by(desc(models.InsumoPuntoVentaPrecio.id))
        .all()
    )

    seen = set()
    unique_results = []
    for item in results:
        if item.punto_venta_id not in seen:
            unique_results.append(item)
            seen.add(item.punto_venta_id)

    return unique_results


def change_insumo_precio_venta_status(
    db: Session, id: int, status: EstadoEnum, modified_by: str
) -> schemas.InsumoPuntoVentaPrecio:
    # Obtener el objeto InsumoPuntoVentaPrecio
    obj = get_insumo_venta_precio_by_id(db, id)

    # Verificar si ya existe un precio activo para el mismo insumo_punto_venta_id
    existing_active_price = db.query(InsumoPuntoVentaPrecio).filter(
        and_(
            InsumoPuntoVentaPrecio.insumo_punto_venta_id == obj.insumo_punto_venta_id,
            InsumoPuntoVentaPrecio.estado == EstadoEnum.ACTIVO.value
        )
    ).first()

    # Si existe un precio activo, evitar cambiar el estado
    if existing_active_price:
        raise HTTPException(
            status_code=400,
            detail="Ya existe un precio activo para este insumo punto de venta."
        )

    # Si no existe un precio activo, proceder con el cambio de estado
    return repositories.change_insumo_venta_precio_status(obj, db, status, modified_by)


def change_inactive_insumo_precio_venta_status(
    db: Session, id: int, status: EstadoEnum, modified_by: str
) -> schemas.InsumoPuntoVentaPrecio:
    obj = get_insumo_venta_precio_by_id(db, id)
    return repositories.change_insumo_venta_precio_status(obj, db, status, modified_by)



def get_insumo_punto_venta_precio_list_by_estado_activo(
    db: Session, gestor_carga_id: int
) -> List[InsumoPuntoVentaPrecio]:
    # Subquery para obtener el precio activo más reciente de cada insumo en cada punto de venta
    subquery = (
        db.query(
            InsumoPuntoVenta.punto_venta_id,
            func.max(InsumoPuntoVentaPrecio.fecha_inicio).label("max_fecha_inicio"),
            InsumoPuntoVentaPrecio.insumo_punto_venta_id
        )
        .join(InsumoPuntoVentaPrecio, InsumoPuntoVenta.id == InsumoPuntoVentaPrecio.insumo_punto_venta_id)
        .filter(
            InsumoPuntoVenta.gestor_carga_id == gestor_carga_id,
            InsumoPuntoVentaPrecio.estado == EstadoEnum.ACTIVO.value
        )
        .group_by(InsumoPuntoVenta.punto_venta_id, InsumoPuntoVentaPrecio.insumo_punto_venta_id)
        .subquery()
    )

    # Consulta principal con DISTINCT ON y ORDER BY correctamente alineado
    return (
        db.query(InsumoPuntoVentaPrecio)
        .join(InsumoPuntoVenta, InsumoPuntoVentaPrecio.insumo_punto_venta_id == InsumoPuntoVenta.id)
        .join(subquery,
              (InsumoPuntoVenta.punto_venta_id == subquery.c.punto_venta_id) &
              (InsumoPuntoVentaPrecio.fecha_inicio == subquery.c.max_fecha_inicio) &
              (InsumoPuntoVentaPrecio.insumo_punto_venta_id == subquery.c.insumo_punto_venta_id))
        .filter(
            InsumoPuntoVenta.gestor_carga_id == gestor_carga_id,
            InsumoPuntoVentaPrecio.estado == EstadoEnum.ACTIVO.value
        )
        .distinct(InsumoPuntoVenta.punto_venta_id)
        .order_by(InsumoPuntoVenta.punto_venta_id, desc(InsumoPuntoVentaPrecio.id))  # Asegurar el ORDER BY coincide con DISTINCT ON
        .all()
    )


def get_all_insumo_punto_venta_precio_list(db: Session) -> List[InsumoPuntoVentaPrecio]:
    return (
        db.query(InsumoPuntoVentaPrecio)
        .join(InsumoPuntoVentaPrecio.insumo_punto_venta)
        .order_by(
            InsumoPuntoVentaPrecio.fecha_inicio,
            # InsumoPuntoVentaPrecio.fecha_fin,
            InsumoPuntoVentaPrecio.modified_by,
        )
        .all()
    )


def get_insumos_by_punto_venta_id_and_gestor_carga(
    db: Session,
    punto_venta_id: int,
    gestor_carga_id: Optional[int],
) -> List[InsumoPuntoVentaPrecio]:
    return (
        db.query(InsumoPuntoVentaPrecio)
        .join(InsumoPuntoVenta, InsumoPuntoVentaPrecio.insumo_punto_venta_id == InsumoPuntoVenta.id)
        .filter(
            InsumoPuntoVenta.punto_venta_id == punto_venta_id,
            InsumoPuntoVenta.gestor_carga_id == gestor_carga_id,
            InsumoPuntoVenta.estado != EstadoEnum.ELIMINADO.value,
            InsumoPuntoVentaPrecio.estado == EstadoEnum.ACTIVO.value
        )
        .order_by(InsumoPuntoVentaPrecio.fecha_inicio.desc())
        .all()
    )

def get_insumo_punto_venta_precio_by_insumo_id_and_moneda_id_and_punto_venta_id(
    db: Session,
    insumo_id: int,
    moneda_id: int,
    punto_venta_id: int,
    gestor_carga_id: Optional[int],
) -> Optional[InsumoPuntoVentaPrecio]:
    insumo_punto_venta = (
        get_insumo_punto_venta_by_insumo_id_and_moneda_id_and_punto_venta_id(
            db,
            insumo_id,
            moneda_id,
            punto_venta_id,
            gestor_carga_id,
        )
    )
    if insumo_punto_venta:
        return get_last_insumo_punto_venta_precio_by_insumo_punto_venta_id(
            db, insumo_punto_venta_id=insumo_punto_venta.id
        )
    return None


def create_insumo_punto_venta_precio(
    db: Session,
    data: InsumoPuntoVentaPrecioForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> InsumoPuntoVentaPrecio:
    gestor_id = get_gestor_carga_by_params(data, gestor_carga_id)
    obj = repositories.create_insumo_punto_venta(
        db,
        data,
        gestor_id,
        modified_by,
    )
    return repositories.create_new_insumo_punto_venta_precio_by_insumo_punto_venta(
        db, obj, data, modified_by
    )


def create_or_update_insumo_punto_venta_precio(
    db: Session,
    data: InsumoPuntoVentaPrecioForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> InsumoPuntoVentaPrecio:
    gestor_id = get_gestor_carga_by_params(data, gestor_carga_id)

    existing_insumo = (
        db.query(InsumoPuntoVenta)
        .filter(
            InsumoPuntoVenta.insumo_id == data.insumo_id,
            InsumoPuntoVenta.punto_venta_id == data.punto_venta_id,
            InsumoPuntoVenta.gestor_carga_id == gestor_id,
            InsumoPuntoVenta.moneda_id == data.moneda_id,
        )
        .first()
    )

    if existing_insumo:
        return update_insumo_punto_venta_precio_by_insumo_punto_venta(
            db, existing_insumo, data, modified_by
        )
    else:
        return crear_precio_insumo_punto_venta(
            db, data, modified_by
        )


def update_insumo_punto_venta_precio_by_insumo_punto_venta(
    db: Session,
    existing_insumo: InsumoPuntoVenta,
    data: InsumoPuntoVentaPrecioForm,
    modified_by: str,
) -> InsumoPuntoVentaPrecio:
    # Obtener el ID del insumo punto de venta
    insumo_punto_venta_id = existing_insumo.id

    # Obtener el precio activo actual para este insumo_punto_venta_id
    current_price = db.query(InsumoPuntoVentaPrecio).filter(
        InsumoPuntoVentaPrecio.insumo_punto_venta_id == insumo_punto_venta_id,
        InsumoPuntoVentaPrecio.estado == EstadoEnum.ACTIVO.value
    ).first()

    # Verificar si el precio ha cambiado
    if current_price and current_price.precio != data.precio:
        # Si el precio cambia, inactivar el precio anterior
        db.query(InsumoPuntoVentaPrecio).filter(
            InsumoPuntoVentaPrecio.insumo_punto_venta_id == insumo_punto_venta_id,
            InsumoPuntoVentaPrecio.estado == EstadoEnum.ACTIVO.value
        ).update({"estado": EstadoEnum.INACTIVO.value}, synchronize_session=False)

        # Crear un nuevo precio con el nuevo valor
        new_price = InsumoPuntoVentaPrecio(
            insumo_punto_venta_id=insumo_punto_venta_id,
            precio=data.precio,
            fecha_inicio=data.fecha_inicio,
            # fecha_fin=data.fecha_fin,
            hora_inicio= data.hora_inicio,
            observacion=data.observacion,
            estado=EstadoEnum.ACTIVO.value,  # El nuevo precio es activo
            created_by=modified_by,
            modified_by=modified_by,
        )
        db.add(new_price)
        db.commit()
        db.refresh(new_price)
        return new_price
    else:
        # Si solo las fechas o la hora cambiaron, actualizamos el registro sin crear uno nuevo
        if current_price:
            current_price.fecha_inicio = data.fecha_inicio
            # current_price.fecha_fin = data.fecha_fin
            current_price.fecha_fin = data.hora_inicio
            current_price.precio = data.precio
            current_price.observacion = data.observacion
            current_price.modified_by = modified_by
            db.commit()
            db.refresh(current_price)
            return current_price
        else:
            # Si no existe un precio, crear un nuevo registro (esto puede no ser necesario según tu lógica)
            new_price = InsumoPuntoVentaPrecio(
                insumo_punto_venta_id=insumo_punto_venta_id,
                precio=data.precio,
                fecha_inicio=data.fecha_inicio,
                # fecha_fin=data.fecha_fin,
                hora_inicio= data.hora_inicio,
                observacion=data.observacion,
                estado=EstadoEnum.ACTIVO.value,
                created_by=modified_by,
                modified_by=modified_by,
            )
            db.add(new_price)
            db.commit()
            db.refresh(new_price)
            return new_price



def update_insumo_punto_venta_precio(
    db: Session,
    id: int,
    data: InsumoPuntoVentaPrecioForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> InsumoPuntoVentaPrecio:
    existing_record = db.query(InsumoPuntoVentaPrecio).filter(InsumoPuntoVentaPrecio.id == id).first()

    if existing_record:
        existing_record.precio = data.precio
        existing_record.fecha_inicio = data.fecha_inicio
        # existing_record.fecha_fin = data.fecha_fin
        existing_record.modified_by = modified_by
        db.commit()
        db.refresh(existing_record)

    return existing_record


def edit_insumo_punto_venta_precio(
    id: int,
    db: Session,
    data: schemas.InsumoPuntoVentaPrecioUpdate,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> InsumoPuntoVentaPrecio:
    gestor_id = get_gestor_carga_by_params(data, gestor_carga_id)
    to_edit_obj = repositories.get_insumo_punto_venta_precio_list_by_id_and_gestor_carga_id(db, id, gestor_id)
    if to_edit_obj:
        to_edit_obj = to_edit_obj[0]  # Selecciona el primer objeto si existe
    else:
        raise HTTPException(status_code=404, detail="InsumoPuntoVentaPrecio no encontrado.")

    return repositories.edit_insumo_punto_venta_precio(to_edit_obj, db, data, modified_by)



def edit_and_create_insumo_punto_venta_precio(
    id: int,
    db: Session,
    data: schemas.InsumoPuntoVentaPrecioUpdate,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> InsumoPuntoVentaPrecio:
    gestor_id = get_gestor_carga_by_params(data, gestor_carga_id)
    to_edit_obj = repositories.get_insumo_punto_venta_precio_list_by_id_and_gestor_carga_id(db, id, gestor_id)

    if not to_edit_obj:
        raise HTTPException(status_code=404, detail="InsumoPuntoVentaPrecio no encontrado.")

    to_edit_obj = to_edit_obj[0]
    if to_edit_obj.precio != data.precio:
        return repositories.create_new_insumo_punto_venta_precio(db, to_edit_obj, data, modified_by)

    return repositories.update_insumo_punto_venta_precio(to_edit_obj, db, data, modified_by)



def get_insumo_punto_venta_precio_by_id(db: Session, id: int) -> InsumoPuntoVentaPrecio:
    obj = repositories.get_insumo_punto_venta_precio_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Insumo no encontrado")
    return obj



def get_insumo_punto_venta_precio_reports(db: Session) -> str:
    datalist = repositories.get_insumo_venta_precio_list(db)
    wb = Workbook()
    ws = wb.active
    i = 0

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "ID"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Precio"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha Inicio"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha Fin"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Usuario creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Usuario modificación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha modificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        i = 0

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.id

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.moneda_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.saldo_confirmado

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.created_by

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.created_at

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.modified_by

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.modified_at

    ws.auto_filter.ref = ws.dimensions
    filename = "insumo_venta_precio_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
