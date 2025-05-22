import os
from typing import List, Optional

from fastapi import HTTPException
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories
from app.config import REPORTS_FOLDER
from app.enums import TipoContraparteEnum
from app.schemas import EstadoCuenta, MovimientoEstadoCuenta, ContraparteEstadoCuenta
from app.logger import logger


def get_estado_cuenta_list(
    db: Session, gestor_carga_id: Optional[int] = None
) -> List[EstadoCuenta]:
    if gestor_carga_id:

        moneda_local= repositories.get_moneda_by_gestor_carga(db, gestor_carga_id)
        moneda_local_id = moneda_local.id if moneda_local else 1

        results = repositories.get_estado_cuenta_list_by_gestor_carga_id(
            db, gestor_carga_id, moneda_local_id
        )

    else:
        results = repositories.get_estado_cuenta_list(db)
    return EstadoCuenta.result_of_query_to_list(results)


def get_estado_cuenta_pdv_list(
    db: Session,
    gestor_carga_id:int,
    tipo_flujo: Optional[str] = None,
    contraparte_id: Optional[int] = None,
    contraparte: Optional[str] = None,
    contraparte_numero_documento: Optional[str] = None,
    punto_venta_id: Optional[int] = None,
) -> List[EstadoCuenta]:

    moneda_local= repositories.get_moneda_by_gestor_carga(db, gestor_carga_id)
    moneda_local_id = moneda_local.id if moneda_local else 1

    results = repositories.get_estado_cuenta_pdv_list(
        db, moneda_local_id, tipo_flujo, contraparte_id, contraparte, contraparte_numero_documento, punto_venta_id
    )

    return EstadoCuenta.result_of_query_to_list(results)


def get_estado_cuenta_pdv(
    db: Session,
    gestor_carga_id:int,
    tipo_flujo: Optional[str] = None,
    contraparte_id: Optional[int] = None,
    contraparte: Optional[str] = None,
    contraparte_numero_documento: Optional[str] = None,
    punto_venta_id: Optional[int] = None,
) -> Optional[EstadoCuenta]:

    moneda_local= repositories.get_moneda_by_gestor_carga(db, gestor_carga_id)
    moneda_local_id = moneda_local.id if moneda_local else 1

    result = repositories.get_estado_cuenta_pdv(
        db, moneda_local_id, tipo_flujo, contraparte_id, contraparte, contraparte_numero_documento, punto_venta_id
    )

    logger.info("result")
    logger.info(result)

    if result:
        retorno = EstadoCuenta.from_orm_row(result)
    else:
        # obtener info contraparte
        result = repositories.get_estado_cuenta_by_contraparte_and_tipo(
            db, contraparte_id, None, punto_venta_id
        )
        retorno = EstadoCuenta.from_orm_row(result)
        retorno.pendiente= 0
        retorno.en_proceso= 0
        retorno.confirmado= 0
        retorno.saldo_pendiente= 0
        retorno.finalizado= 0
        retorno.cantidad_pendiente= 0
        retorno.cantidad_en_proceso= 0
        retorno.cantidad_confirmado= 0
        retorno.cantidad_finalizado= 0

    return retorno


def get_estado_cuenta_by_contraparte(
    db: Session,
    tipo_contraparte_id: int,
    contraparte_id: int,
    contraparte: str,
    contraparte_numero_documento: str,
    punto_venta_id: int = None,
) -> Optional[EstadoCuenta]:
    tipo_contraparte = repositories.get_tipo_comprobante_by_id(db, tipo_contraparte_id)
    # obtenemos la mon local
    if not tipo_contraparte:
        raise HTTPException(status_code=404, detail="Tipo de Contraparte no encontrado")
    if tipo_contraparte.descripcion == TipoContraparteEnum.OTRO.value:
        result = repositories.get_estado_cuenta_by_contraparte_tipo_otro(
            db, contraparte, contraparte_numero_documento, tipo_contraparte_id
        )
    else:
        result = repositories.get_estado_cuenta_by_contraparte_and_tipo(
            db, contraparte_id, 1, tipo_contraparte_id, punto_venta_id
        )
    if result:
        return EstadoCuenta.from_orm_row(result)
    return None


def get_estado_cuenta_reports(db: Session) -> str:
    datalist = get_estado_cuenta_list(db)
    wb = Workbook()
    ws = wb.active

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Tipo de Contraparte"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Nombre Contraparte"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "Nº Contraparte"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Pendiente"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "En Proceso"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=6)
    title_cell.value = "Confirmado"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Finalizado"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.tipo_contraparte_descripcion

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.contraparte

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.contraparte_numero_documento

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.pendiente

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = item.en_proceso

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.confirmado

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = item.finalizado

    ws.auto_filter.ref = ws.dimensions
    filename = "estado_cuenta_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename


def get_nuevo_servicio(
    db: Session,
    tipo_contraparte_id: int,
    contraparte_id: int,
    contraparte: str,
    contraparte_numero_documento: str,
    gestor_carga_id: Optional[int],
    punto_venta_id: Optional[int] = None,
    linea_movimiento: Optional[str] = None,
) -> List[MovimientoEstadoCuenta]:

    results = repositories.nuevo_endpint(
        db, tipo_contraparte_id, contraparte_id, contraparte,
        contraparte_numero_documento, gestor_carga_id, punto_venta_id,
        linea_movimiento
    )

    return MovimientoEstadoCuenta.result_of_query_to_list(results)


def get_saldo_cuenta_contraparte(
    db: Session,
    gestor_carga_id: Optional[int],
    tipo_contraparte_id: int,
    contraparte_id: int,
    punto_venta_id: Optional[int] = None,
) -> ContraparteEstadoCuenta :

    results = repositories.get_saldo_cuenta_contraparte(
            db, gestor_carga_id, tipo_contraparte_id, contraparte_id, punto_venta_id
        )

    return ContraparteEstadoCuenta.from_orm_row(results)


def get_report_nuevo_servicio(
    db: Session,
    tipo_contraparte_id: int,
    contraparte_id: int,
    contraparte: str,
    contraparte_numero_documento: str,
    gestor_carga_id: Optional[int],
    punto_venta_id: Optional[int] = None,
    linea_movimiento: Optional[str] = None,
) -> str:
    results = []

    if gestor_carga_id:
        results = repositories.nuevo_endpint(
            db, tipo_contraparte_id, contraparte_id, contraparte,
            contraparte_numero_documento, gestor_carga_id, punto_venta_id, linea_movimiento
        )

    return get_movimiento_estado_cuenta_reports(results)


def get_movimiento_estado_cuenta_reports(datalist: List[MovimientoEstadoCuenta],) -> str:

    wb = Workbook()
    ws = wb.active
    i = 0

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Nº de Movimiento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Estado"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Cuenta"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Concepto"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Detalle"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "N°OC"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Info"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "N°Liq"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Pendiente"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Confirmado"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Pago/Cobro"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        i = 0

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.movimiento_id

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.estado

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.fecha

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.tipo_cuenta_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.tipo_movimiento_concepto

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.detalle

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.nro_documento_relacionado

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.info

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.liquidacion_id

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.estado_liquidacion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.pendiente

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.confirmado

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.finalizado

    ws.auto_filter.ref = ws.dimensions
    filename = "estado_cuenta_movimiento_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))

    return filename
