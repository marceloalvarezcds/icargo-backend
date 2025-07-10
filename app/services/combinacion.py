import asyncio
from datetime import datetime
import os
from typing import List, Optional

from app.config import REPORTS_FOLDER
from app.enums.estado import EstadoEnum
from app.models.camion import Camion
from app.models.chofer import Chofer
from app.models.orden_carga import OrdenCarga
from app.models.semi import Semi
from app.utils.gestor_carga import get_gestor_carga_by_params
from fastapi import HTTPException # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore
from app.models import Combinacion
from app import repositories, schemas
from .semi import get_semi_by_id
from http import HTTPStatus
from fastapi import HTTPException

def get_combinacion_list(
    db: Session, gestor_carga_id: Optional[int]
) -> List[Combinacion]:
    if gestor_carga_id:
        return repositories.get_combinacion_list_by_gestor_carga_id(db, gestor_carga_id)
    return repositories.get_combinacion_list(db)


def get_combinacion_all_list(
    db: Session, gestor_carga_id: Optional[int]
) -> List[Combinacion]:
    if gestor_carga_id:
        return repositories.get_combinacion_all_list_by_gestor_carga_id(db, gestor_carga_id)
    return repositories.get_combinacion_list(db)


def get_combinacion_activa_list(
    db: Session, gestor_carga_id: Optional[int]
) -> List[Combinacion]:
    if gestor_carga_id:
        return repositories.get_combinacion_activa_list_by_gestor_carga_id(db, gestor_carga_id)
    return repositories.get_combinacion_list_combinacion_activa(db)


def get_combinacion_by_id(db: Session, id: int) -> Combinacion:
    obj = repositories.get_combinacion_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Combinacion no encontrada")
    return obj


def get_combinacion_by_gestor_cuenta_and_combinacion_id(
    db: Session, semi_id: int, gestor_cuenta_id: Optional[int]
) -> List[Combinacion]:
    semi = get_semi_by_id(db, semi_id)
    lista = repositories.get_propietario_list_by_gestor_cuenta_id(db, gestor_cuenta_id)
    if semi.propietario:
        combinacion: Combinacion = semi.combinacion
        lista.append(combinacion)
    return lista


def change_combinacion_status(
    db: Session, id: int, status: EstadoEnum, modified_by: str
) -> schemas.CombinacionesBD:
    co = get_combinacion_by_id(db, id)
    return repositories.change_combinacion_status(co, db, status, modified_by)


async def rol_tiene_permiso_cambiar_estado(
    db: Session,
    rol_id: int
) -> bool:

    gestor_carga_rol = repositories.get_rol_list(db, rol_id)
    if not gestor_carga_rol:
        return False

    for permiso in gestor_carga_rol.permisos:
        if permiso.accion == "cambiar_estado":
            return True

    return False


def get_camion_by_combinacion_id(
    db: Session, gestor_carga_id: Optional[int]
) -> List[Combinacion]:
    camion_semi_neto_list = repositories.get_camion_list_by_combinacion_id(
        db,  gestor_carga_id
    )
    return camion_semi_neto_list


def get_camion_list_combinacion(
    db: Session, gestor_carga_id: Optional[int]
) -> List[Camion]:
    id_list: List[int] = []
    filtered_list: List[Camion] = []
    original_list = get_camion_by_combinacion_id(
        db, gestor_carga_id
    )
    camion_list: List[Camion] = list(map(lambda x: x.camion, original_list))
    for item in camion_list:
        # check if exists in unique_list or not, and is estado ACTIVO
        chofer: Optional[Chofer] = item.chofer
        if (
            item.id not in id_list
            and item.estado == EstadoEnum.ACTIVO.value
            and (chofer and chofer.estado == EstadoEnum.ACTIVO.value)
        ):
            id_list.append(item.id)
            filtered_list.append(item)
    return filtered_list


def get_combinacion_semi_list_by_camion_id(
    db: Session, camion_id: int, gestor_carga_id: int
) -> List[Combinacion]:
    camion_semi_neto_list = (
        repositories.get_semi_list_by_camion_id(
            db, camion_id, gestor_carga_id
        )
    )
    return camion_semi_neto_list


def get_semi_list_by_camion_id(
    db: Session, camion_id: int, gestor_carga_id: Optional[int]
) -> List[Semi]:
    combinaciones = (
        db.query(Combinacion)
        .filter(
            Combinacion.camion_id == camion_id,
            Combinacion.gestor_carga_id == gestor_carga_id if gestor_carga_id is not None else Combinacion.gestor_carga_id != None,
            Combinacion.estado == EstadoEnum.ACTIVO.value
        )
        .all()
    )

    semi_list: List[Semi] = [combinacion.semi for combinacion in combinaciones]

    return semi_list


def get_combinacion_by_camion_id_and_semi_id_(
    db: Session,
    camion_id: int,
    semi_id: int,
    gestor_carga_id: Optional[int],
) -> Optional[Combinacion]:
    obj = None
    obj = (
        repositories.get_camion_id_and_semi_id(
            db, camion_id, semi_id, gestor_carga_id
        )
    )
    if not obj:
        obj = repositories.get_camion_id_and_semi_id(
            db, camion_id, semi_id, gestor_carga_id
        )
    return obj


async def create_combinacion(
    db: Session,
    data: schemas.CombinacionCreateModel,
    modified_by: str,
    gestor_carga_id: Optional[int],
) -> schemas.Combinacion:
    propietario = repositories.get_propietario_by_id(db, data.propietario_id)
    camion = repositories.get_camion_by_id(db, data.camion_id)
    chofer = repositories.get_chofer_by_id(db, data.chofer_id)
    semi = repositories.get_semi_by_id(db, data.semi_id)

    if not propietario:
        raise HTTPException(
            status_code=404,
            detail="El propietario especificado no existe."
        )
    if not camion:
        raise HTTPException(
            status_code=404,
            detail="El camión especificado no existe."
        )
    if not chofer:
        raise HTTPException(
            status_code=404,
            detail="El chofer especificado no existe."
        )
    if not semi:
        raise HTTPException(
            status_code=404,
            detail="El semi especificado no existe."
        )

    if propietario.estado != EstadoEnum.ACTIVO.value:
        raise HTTPException(
            status_code=HTTPStatus.BAD_REQUEST,
            detail=f"El Propietario '{propietario.nombre}' está INACTIVO. No se puede crear la Combinación."
        )
    if camion.estado != EstadoEnum.ACTIVO.value:
        raise HTTPException(
            status_code=HTTPStatus.BAD_REQUEST,
            detail=f"El Tracto con placa '{camion.placa}' está INACTIVO. No se puede crear la Combinación."
        )
    if chofer.estado != EstadoEnum.ACTIVO.value:
        raise HTTPException(
            status_code=HTTPStatus.BAD_REQUEST,
            detail=f"El Chofer '{chofer.nombre}' está INACTIVO. No se puede crear la Combinación."
        )
    if semi.estado != EstadoEnum.ACTIVO.value:
        raise HTTPException(
            status_code=HTTPStatus.BAD_REQUEST,
            detail=f"El Semi con placa '{semi.placa}' está INACTIVO. No se puede crear la Combinación."
        )

    combinacion_tracto_propietario = repositories.get_combinacion_tracto_propietario_ids(
        db, data.camion_id, data.propietario_id, gestor_carga_id
    )
    if combinacion_tracto_propietario:
        raise HTTPException(
            status_code=409,
            detail="El tracto puede tener un solo beneficiario."
        )

    combinacion = repositories.create_combinacion(
        db,
        data,
        gestor_carga_id,
        modified_by,
    )
    return combinacion


def edit_combinacion(
    id: int,
    db: Session,
    data: schemas.CombinacionUpdate,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Combinacion:
    combinacion_to_edit = db.query(Combinacion).filter(Combinacion.id == id).first()
    if not combinacion_to_edit:
        raise HTTPException(status_code=404, detail="Combinación no encontrada")

    # Verificar combinaciones similares
    combinacion_exists = repositories.get_combinacion_by_ids(
        db, data.propietario_id, data.camion_id, data.chofer_id, gestor_carga_id
    )
    combinacion_tracto = repositories.get_combinacion_tracto_ids(
        db, data.camion_id, gestor_carga_id
    )

    # Permitir editar límites sin validar conflictos de combinación
    solo_limites_modificados = (
        data.limite_monto_anticipos is not None or
        data.camion_oc_activa is not None
    ) and not (
        data.propietario_id != combinacion_to_edit.propietario_id or
        data.camion_id != combinacion_to_edit.camion_id or
        data.chofer_id != combinacion_to_edit.chofer_id or
        data.semi_id != combinacion_to_edit.semi_id
    )

    if not solo_limites_modificados:
        # Validaciones solo si cambian camión, propietario o chofer
        if (data.propietario_id != combinacion_to_edit.propietario_id or
            data.camion_id != combinacion_to_edit.camion_id):

            combinacion_tracto_propietario = db.query(Combinacion).filter(
                Combinacion.camion_id == data.camion_id,
                Combinacion.propietario_id == data.propietario_id,
                Combinacion.estado != EstadoEnum.INACTIVO.value
            ).first()

            if combinacion_tracto_propietario:
                raise HTTPException(
                    status_code=409,
                    detail="El tracto ya tiene un beneficiario asociado."
                )
    return repositories.edit_combinacion(combinacion_to_edit, db, data, gestor_carga_id, modified_by)


def get_combinacion_reports(db: Session) -> str:
    datalist = repositories.get_combinacion_list(db)
    wb = Workbook()
    # get worksheet
    ws = wb.active


    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Estado"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Comentario"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=8)
    title_cell.value = "Capacidad Total Combinación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=9)
    title_cell.value = "Usuario creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=10)
    title_cell.value = "Fecha creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=11)
    title_cell.value = "Usuario modificación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=12)
    title_cell.value = "Fecha modificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.estado

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.comentario

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.capacidad_total_combinacion

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = item.created_by

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.created_at

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = item.modified_by

        value_cell = ws.cell(row=row + 2, column=8)
        value_cell.value = item.modified_at

    ws.auto_filter.ref = ws.dimensions
    filename = "combinacion_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
