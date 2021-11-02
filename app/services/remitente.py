import os
from typing import List, Optional

from fastapi import HTTPException, UploadFile  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import logger, repositories, schemas
from app.config import REPORTS_FOLDER
from app.models import GestorCargaRemitente, Remitente, RemitenteContactoGestorCarga

from .gestor_carga_remitente import (
    create_gestor_carga_remitente,
    edit_gestor_carga_remitente,
)
from .pictshare import upload_and_get_image_url
from .remitente_contacto import update_remitente_contacto_list


def get_remitente_detail(
    obj: Remitente, gestor_carga_id: Optional[int]
) -> schemas.Remitente:
    obj_dict = obj.as_dict(for_json=False)
    contactos: List[RemitenteContactoGestorCarga] = obj.contactos
    ges: List[GestorCargaRemitente] = obj.gestores
    gestores = [x for x in ges if x.gestor_carga_id == gestor_carga_id]
    obj_dict["contactos"] = [
        x for x in contactos if x.gestor_carga_id == gestor_carga_id
    ]
    obj_dict["gestor_carga_remitente"] = gestores[0] if len(gestores) > 0 else None
    obj_dict["tipo_documento"] = obj.tipo_documento
    obj_dict["composicion_juridica"] = obj.composicion_juridica
    obj_dict["ciudad"] = obj.ciudad
    logger.info(obj_dict)
    return schemas.Remitente.parse_obj(obj_dict)


async def create_remitente(
    db: Session,
    data: schemas.RemitenteForm,
    file: UploadFile,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> schemas.Remitente:
    logo_url = await upload_and_get_image_url(file)
    obj = repositories.create_remitente(db, data, logo_url, modified_by)
    update_remitente_contacto_list(
        db, data.contactos, obj, gestor_carga_id, modified_by
    )
    create_gestor_carga_remitente(db, obj, gestor_carga_id, data.alias, modified_by)
    return get_remitente_detail(obj, gestor_carga_id)


def get_remitente_by_id(db: Session, id: int) -> Remitente:
    obj = repositories.get_remitente_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Centro Operativo no encontrado")
    return obj


def get_remitente_by_id_and_gestor_carga_id(
    db: Session, id: int, gestor_carga_id: Optional[int] = None
) -> schemas.Remitente:
    obj = repositories.get_remitente_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Centro Operativo no encontrado")
    return get_remitente_detail(obj, gestor_carga_id)


async def edit_remitente(
    id: int,
    db: Session,
    data: schemas.RemitenteForm,
    file: Optional[UploadFile],
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> schemas.Remitente:
    logger.info(data)
    logo_url = await upload_and_get_image_url(file) if file else None
    to_edit_obj = get_remitente_by_id(db, id)
    obj = repositories.edit_remitente(to_edit_obj, db, data, logo_url, modified_by)
    update_remitente_contacto_list(
        db, data.contactos, obj, gestor_carga_id, modified_by
    )
    edit_gestor_carga_remitente(db, obj, gestor_carga_id, data.alias, modified_by)
    return get_remitente_detail(obj, gestor_carga_id)


def delete_remitente(
    db: Session, id: int, gestor_carga_id: Optional[int], modified_by: str
) -> schemas.Remitente:
    co = get_remitente_by_id(db, id)
    obj = repositories.delete_remitente(co, db, modified_by)
    return get_remitente_detail(obj, gestor_carga_id)


def get_remitente_reports(db: Session) -> str:
    datalist = repositories.get_remitente_list(db)
    wb = Workbook()
    # get worksheet
    ws = wb.active
    # title = f"Lista de Centros Operativos"
    # title_cell = ws.cell(row=1, column=1)
    # title_cell.font = Font(size=12, bold=True)
    # title_cell.value = title.upper()

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Nombre"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Nombre de Fantasía"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "Tipo de Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Número de Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Composición Jurídica"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=6)
    title_cell.value = "Dirección"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Ubicación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.nombre

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.nombre_corto if item.nombre_corto else ""

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.tipo_documento.descripcion

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.numero_documento

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = (
            item.composicion_juridica.nombre if item.composicion_juridica else ""
        )

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.direccion if item.direccion else ""

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = f"{item.ciudad.nombre}/{item.ciudad.localidad.nombre}/{item.ciudad.localidad.pais.nombre_corto}"  # noqa

    ws.auto_filter.ref = ws.dimensions
    filename = "remitente_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
