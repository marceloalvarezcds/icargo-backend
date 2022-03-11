import os
from typing import List, Optional

from fastapi import HTTPException  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories
from app.config import REPORTS_FOLDER
from app.models import Caja
from app.schemas import CajaForm


def get_caja_list(db: Session, gestor_carga_id: Optional[int]) -> List[Caja]:
    if gestor_carga_id:
        return repositories.get_caja_list_by_gestor_carga_id(db, gestor_carga_id)
    return repositories.get_caja_list(db)


def create_caja(
    db: Session,
    data: CajaForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Caja:
    gestor_id = gestor_carga_id if gestor_carga_id else data.gestor_carga_id
    if not gestor_id:
        raise HTTPException(status_code=409, detail="Debe elegir un Gestor de carga")
    if repositories.get_caja_by(db, data.nombre, gestor_id):
        raise HTTPException(status_code=409, detail=f"La Caja {data.nombre} ya existe")
    return repositories.create_caja(db, data, gestor_id, modified_by)


def get_caja_by_id(db: Session, id: int) -> Caja:
    obj = repositories.get_caja_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Caja no encontrada")
    return obj


def edit_caja(
    id: int,
    db: Session,
    data: CajaForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Caja:
    gestor_id = gestor_carga_id if gestor_carga_id else data.gestor_carga_id
    if not gestor_id:
        raise HTTPException(status_code=409, detail="Debe elegir un Gestor de carga")
    exists = repositories.get_caja_by(db, data.nombre, gestor_id)
    if exists and exists.id != id:
        raise HTTPException(status_code=409, detail=f"La Caja {data.nombre} ya existe")
    to_edit_obj = get_caja_by_id(db, id)
    return repositories.edit_caja(to_edit_obj, db, data, gestor_id, modified_by)


def delete_caja(db: Session, id: int, modified_by: str) -> Caja:
    co = get_caja_by_id(db, id)
    return repositories.delete_caja(co, db, modified_by)


def get_caja_reports(db: Session) -> str:
    datalist = repositories.get_caja_list(db)
    wb = Workbook()
    ws = wb.active

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "ID"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Nombre"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "Moneda"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Crédito"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Débito"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=6)
    title_cell.value = "Saldo"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Usuario creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=8)
    title_cell.value = "Fecha creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=9)
    title_cell.value = "Usuario modificación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=10)
    title_cell.value = "Fecha modificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.id

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.nombre

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.moneda_nombre

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.credito

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = item.debito

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.saldo_confirmado

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = item.created_by

        value_cell = ws.cell(row=row + 2, column=8)
        value_cell.value = item.created_at

        value_cell = ws.cell(row=row + 2, column=9)
        value_cell.value = item.modified_by

        value_cell = ws.cell(row=row + 2, column=10)
        value_cell.value = item.modified_at

    ws.auto_filter.ref = ws.dimensions
    filename = "caja_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
