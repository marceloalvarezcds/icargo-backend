import os
from typing import Optional

from fastapi import HTTPException, UploadFile  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories, schemas
from app.config import REPORTS_FOLDER
from app.enums.estado import EstadoEnum
from app.models import GestorCarga

from .pictshare import upload_and_get_image_url


async def create_gestor_carga(
    db: Session,
    data: schemas.GestorCargaForm,
    file: Optional[UploadFile],
    modified_by: str,
) -> schemas.GestorCarga:
    if repositories.get_gestor_carga_by(
        db, data.tipo_documento_id, data.numero_documento
    ):
        raise HTTPException(
            status_code=409,
            detail=f"El Gestor de Carga con documento {data.numero_documento} ya existe",
        )
    logo_url = await upload_and_get_image_url(file)
    return repositories.create_gestor_carga(db, data, logo_url, modified_by)


def get_gestor_carga_by_id(db: Session, id: int) -> GestorCarga:
    obj = repositories.get_gestor_carga_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Gestor de Carga no encontrado")
    return obj


async def edit_gestor_carga(
    id: int,
    db: Session,
    data: schemas.GestorCargaForm,
    file: Optional[UploadFile],
    modified_by: str,
) -> schemas.GestorCarga:
    exists = repositories.get_gestor_carga_by(
        db, data.tipo_documento_id, data.numero_documento
    )
    if exists and exists.id != id:
        raise HTTPException(
            status_code=409,
            detail=f"El Gestor de Carga con documento {data.numero_documento} ya existe",
        )
    logo_url = await upload_and_get_image_url(file) if file else None
    to_edit_obj = get_gestor_carga_by_id(db, id)
    return repositories.edit_gestor_carga(to_edit_obj, db, data, logo_url, modified_by)


def delete_gestor_carga(db: Session, id: int, modified_by: str) -> schemas.GestorCarga:
    co = get_gestor_carga_by_id(db, id)
    return repositories.delete_gestor_carga(co, db, modified_by)


def change_gestor_carga_status(
    db: Session, id: int, status: EstadoEnum, modified_by: str
) -> schemas.CentroOperativo:
    co = get_gestor_carga_by_id(db, id)
    return repositories.change_gestor_carga_status(co, db, status, modified_by)


def get_gestor_carga_reports(db: Session) -> str:
    datalist = repositories.get_gestor_carga_list(db)
    wb = Workbook()
    # get worksheet
    ws = wb.active

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Nombre"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Nombre de Fantasía"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "Tipo de Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Número de Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Composición Jurídica"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=6)
    title_cell.value = "Dirección"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Ubicación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.nombre

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.nombre_corto if item.nombre_corto else ""

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.tipo_documento.descripcion

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.numero_documento

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = (
            item.composicion_juridica.nombre if item.composicion_juridica else ""
        )

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.direccion if item.direccion else ""

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = (
            f"{item.ciudad.nombre}/{item.ciudad.localidad.nombre}/{item.ciudad.localidad.pais.nombre_corto}"  # noqa: B950
            if item.ciudad
            else ""
        )

    ws.auto_filter.ref = ws.dimensions
    filename = "gestor_carga_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
