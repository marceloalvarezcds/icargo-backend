import os
from typing import List, Optional, cast

from fastapi import HTTPException, UploadFile  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories, schemas
from app.config import REPORTS_FOLDER
from app.enums import EstadoEnum
from app.models import Chofer
from app.models.combinacion import Combinacion
from app.utils import get_gestor_carga_by_params

from .camion import get_camion_by_id
from .chofer_anticipos_oc import bloquear_anticipos_desde_el_chofer
from .chofer_check_files import check_files, get_chofer_detail
from .chofer_propietario import (
    create_or_edit_propietario_by_chofer,
    disable_propietario_by_ruc,
)
from .gestor_carga_chofer import create_gestor_carga_chofer, edit_gestor_carga_chofer


def get_chofer_list_without_camion_by_camion_id(
    db: Session, camion_id: int, gestor_cuenta_id: Optional[int]
) -> List[Chofer]:
    camion = get_camion_by_id(db, camion_id)
    lista = repositories.get_chofer_list_without_camion(db, gestor_cuenta_id)
    if camion.chofer:
        chofer: Chofer = camion.chofer
        lista.append(chofer)
    return lista


async def create_chofer(
    db: Session,
    data: schemas.ChoferForm,
    foto_documento_frente_file: Optional[UploadFile],
    foto_documento_reverso_file: Optional[UploadFile],
    foto_perfil_file: Optional[UploadFile],
    foto_registro_frente_file: Optional[UploadFile],
    foto_registro_reverso_file: Optional[UploadFile],
    foto_documento_frente_propietario_file: Optional[UploadFile],
    foto_documento_reverso_propietario_file: Optional[UploadFile],
    gestor_cuenta_id: Optional[int],
    modified_by: str,
    usuario_id: int,
) -> schemas.Chofer:
    if repositories.get_chofer_by(
        db, data.tipo_documento_id, data.pais_emisor_documento_id, data.numero_documento
    ):
        raise HTTPException(
            status_code=409,
            detail=f"El Chofer con documento {data.numero_documento} ya existe",
        )
    (
        foto_documento_frente_url,
        foto_documento_reverso_url,
        foto_registro_reverso_url,
        foto_registro_frente_url,
        foto_perfil_url,
    ) = await check_files(
        foto_documento_frente_file,
        foto_documento_reverso_file,
        foto_registro_frente_file,
        foto_registro_reverso_file,
        foto_perfil_file,
    )
    obj = repositories.create_chofer(
        db,
        data,
        gestor_cuenta_id,
        foto_documento_frente_url,
        foto_documento_reverso_url,
        foto_perfil_url,
        foto_registro_frente_url,
        foto_registro_reverso_url,
        modified_by,
        usuario_id,
    )
    create_gestor_carga_chofer(db, obj, gestor_cuenta_id, data.alias, modified_by)
    if data.es_propietario:
        await create_or_edit_propietario_by_chofer(
            db,
            data,
            foto_documento_frente_propietario_file,
            foto_documento_reverso_propietario_file,
            foto_perfil_url,
            gestor_cuenta_id,
            obj,
            modified_by,
        )
    return get_chofer_detail(db, obj, gestor_cuenta_id)


def get_chofer_by_id(db: Session, id: int) -> Chofer:
    obj = repositories.get_chofer_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Chofer no encontrado")
    return obj


def get_chofer_by_id_and_gestor_cuenta_id(
    db: Session, id: int, gestor_cuenta_id: Optional[int] = None
) -> schemas.Chofer:
    obj = repositories.get_chofer_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Chofer no encontrado")
    return get_chofer_detail(db, obj, gestor_cuenta_id)


async def edit_chofer(
    id: int,
    db: Session,
    data: schemas.ChoferEditForm,
    foto_documento_frente_file: Optional[UploadFile],
    foto_documento_reverso_file: Optional[UploadFile],
    foto_perfil_file: Optional[UploadFile],
    foto_registro_frente_file: Optional[UploadFile],
    foto_registro_reverso_file: Optional[UploadFile],
    foto_documento_frente_propietario_file: Optional[UploadFile],
    foto_documento_reverso_propietario_file: Optional[UploadFile],
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> schemas.Chofer:
    gestor_id = get_gestor_carga_by_params(data, gestor_carga_id)
    if (
        data.tipo_documento_id
        and data.pais_emisor_documento_id
        and data.numero_documento
    ):
        exists = repositories.get_chofer_by(
            db,
            data.tipo_documento_id,
            data.pais_emisor_documento_id,
            data.numero_documento,
        )
        if exists and exists.id != id:
            raise HTTPException(
                status_code=409,
                detail=f"El Chofer con documento {data.numero_documento} ya existe",
            )
    (
        foto_documento_frente_url,
        foto_documento_reverso_url,
        foto_registro_reverso_url,
        foto_registro_frente_url,
        foto_perfil_url,
    ) = await check_files(
        foto_documento_frente_file,
        foto_documento_reverso_file,
        foto_registro_frente_file,
        foto_registro_reverso_file,
        foto_perfil_file,
    )
    to_edit_obj = get_chofer_by_id(db, id)
    obj = repositories.edit_chofer(
        to_edit_obj,
        db,
        data,
        foto_documento_frente_url,
        foto_documento_reverso_url,
        foto_perfil_url,
        foto_registro_frente_url,
        foto_registro_reverso_url,
        modified_by,
    )
    edit_gestor_carga_chofer(db, obj, gestor_id, data.alias, modified_by)
    if data.es_propietario:
        chofer_data = cast(schemas.ChoferForm, data)
        await create_or_edit_propietario_by_chofer(
            db,
            chofer_data,
            foto_documento_frente_propietario_file,
            foto_documento_reverso_propietario_file,
            foto_perfil_url,
            gestor_id,
            obj,
            modified_by,
        )
    else:
        disable_propietario_by_ruc(db, to_edit_obj.ruc, modified_by)
    bloquear_anticipos_desde_el_chofer(db, id, data, gestor_id, modified_by)
    return get_chofer_detail(db, obj, gestor_id)


def delete_chofer(
    db: Session, id: int, gestor_cuenta_id: Optional[int], modified_by: str
) -> schemas.Chofer:
    co = get_chofer_by_id(db, id)
    obj = repositories.delete_chofer(co, db, modified_by)
    return get_chofer_detail(db, obj, gestor_cuenta_id)


def change_chofer_status(
    db: Session, id: int, status: EstadoEnum, modified_by: str
) -> schemas.Chofer:
    # Obtener el chofer
    chofer = get_chofer_by_id(db, id)
    if not chofer:
        raise HTTPException(status_code=404, detail="Chofer no encontrado.")

    repositories.change_chofer_status(chofer, db, status, modified_by)
    combinaciones_relacionadas = repositories.get_combinaciones_by_chofer_id(db, chofer.id)

    if status == EstadoEnum.INACTIVO:
        # Inactivar cada combinación asociada
        for combinacion in combinaciones_relacionadas:
            repositories.change_combinacion_status(combinacion, db, EstadoEnum.INACTIVO, modified_by)

    elif status == EstadoEnum.ACTIVO:
        for combinacion in combinaciones_relacionadas:
            # Verificar que todos los elementos estén activos
            if (
                combinacion.camion and combinacion.camion.estado == EstadoEnum.ACTIVO.value and
                combinacion.semi and combinacion.semi.estado == EstadoEnum.ACTIVO.value and
                combinacion.propietario and combinacion.propietario.estado == EstadoEnum.ACTIVO.value
            ):
                # Verificar que no exista otra combinación activa con el mismo camion y propietario
                conflicto = db.query(Combinacion).filter(
                    Combinacion.id != combinacion.id,
                    Combinacion.camion_id == combinacion.camion_id,
                    Combinacion.propietario_id == combinacion.propietario_id,
                    Combinacion.estado != EstadoEnum.INACTIVO.value
                ).first()

                if not conflicto:
                    repositories.change_combinacion_status(combinacion, db, EstadoEnum.ACTIVO, modified_by)
                # Si hay conflicto, no se activa la combinación
    db.commit()

    return schemas.Chofer.from_orm(chofer)


def get_chofer_reports(db: Session) -> str:
    datalist = repositories.get_chofer_list(db)
    wb = Workbook()
    # get worksheet
    ws = wb.active

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Nombre"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Tipo de Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "País Emisor del Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Número de Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Fecha de Nacimiento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=6)
    title_cell.value = "Gestor de Cuenta"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Oficial de Cuenta"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=8)
    title_cell.value = "Dirección"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=9)
    title_cell.value = "Ubicación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=10)
    title_cell.value = "Usuario creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=11)
    title_cell.value = "Fecha creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=12)
    title_cell.value = "Usuario modificación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=13)
    title_cell.value = "Fecha modificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.nombre

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.tipo_documento.descripcion

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = (
            item.pais_emisor_documento.nombre if item.pais_emisor_documento else None
        )

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.numero_documento

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = item.fecha_nacimiento

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.gestor_cuenta_nombre

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = item.oficial_cuenta_nombre

        value_cell = ws.cell(row=row + 2, column=8)
        value_cell.value = item.direccion if item.direccion else ""

        value_cell = ws.cell(row=row + 2, column=9)
        value_cell.value = (
            f"{item.ciudad.nombre}/{item.ciudad.localidad.nombre}/{item.ciudad.localidad.pais.nombre_corto}"  # noqa: B950
            if item.ciudad
            else ""
        )

        value_cell = ws.cell(row=row + 2, column=10)
        value_cell.value = item.created_by

        value_cell = ws.cell(row=row + 2, column=11)
        value_cell.value = item.created_at

        value_cell = ws.cell(row=row + 2, column=12)
        value_cell.value = item.modified_by

        value_cell = ws.cell(row=row + 2, column=13)
        value_cell.value = item.modified_at

    ws.auto_filter.ref = ws.dimensions
    filename = "chofer_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
