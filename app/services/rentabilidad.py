import os
from typing import List, Optional

from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories
from app.config import REPORTS_FOLDER
from app.schemas import Rentabilidad


def get_rentabilidad_list(
    db: Session, gestor_carga_id: Optional[int]
) -> List[Rentabilidad]:
    lista = repositories.get_orden_carga_list(db)
    if gestor_carga_id:
        lista = repositories.get_orden_carga_list_by_gestor_carga_id(
            db, gestor_carga_id
        )
    return Rentabilidad.get_list_by_oc(lista)


def get_rentabilidad_reports(db: Session, gestor_carga_id: Optional[int]) -> str:
    datalist = get_rentabilidad_list(db, gestor_carga_id)
    wb = Workbook()
    # get worksheet
    ws = wb.active
    i = 0

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Estado OC"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Nº OC"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Nº Pedido"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Nº Doc Remisiones"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Estado Anticipos"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Chofer"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Chofer: Tipo Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Chofer: Nº del Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Placa Camión"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Placa Semi"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Propietario"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tipo de flete"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Producto"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Cant. Nominada (kg)"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tot. Cant. Origen"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tot. Cant. Destino"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Diferencia"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Remitente"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Origen"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Destino"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Lugar de Carga"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Lugar de Descarga"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tarifa Flete a PAGAR"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Moneda de Pago"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Unidad de Pago"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tot. Flete a Pagar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tot. Flete a Pagar (Moneda Local)"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Precio Merma a Cobrar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Moneda Merma a Cobrar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Unidad Merma a Cobrar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tolerancia Propietario"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Merma Propietario Kg"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Valor de Merma a Cobrar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tarifa Flete a COBRAR"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Moneda de Cobro"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Unidad de Cobro"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tot. Flete a Cobrar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tot. Flete a Cobrar (Moneda Local)"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Precio Merma a PAGAR"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Moneda Merma a Pagar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Unidad Merma a Pagar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tolerancia Gestora de Carga"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Merma Gestora Carga Kg"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Valor de Merma a Pagar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Total Complemento a Pagar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Total Complemento a Cobrar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Diferencia Complemento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Total Descuento a Pagar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Total Descuento a Cobrar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Diferencia Descuento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tot. Anticipos retirados"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Resultado Gestora de Carga (ML)"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Saldo Final Propietario"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha Conciliación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Usuario creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Usuario modificación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha modificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        i = 0
        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.estado

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.oc_id

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.flete_id

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.nro_remisiones

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.estado_anticipo

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.chofer_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.chofer_tipo_documento

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.chofer_numero_documento

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.camion_placa

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.semi_placa

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.propietario_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.flete_tipo if item.flete_tipo else ""

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.producto_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.cantidad_nominada

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.cantidad_origen

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.cantidad_destino

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.diferencia_origen_destino

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.remitente_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.origen_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.destino_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.lugar_carga_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.lugar_descarga_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_propietario_tarifa

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_propietario_moneda_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_propietario_unidad_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.propietario_flete_total

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.propietario_flete_total_ml

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_propietario_valor

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_propietario_moneda_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_propietario_unidad_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_propietario_tolerancia

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_propietario_merma

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_propietario_valor_merma

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_gestor_carga_tarifa

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_gestor_carga_moneda_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_gestor_carga_unidad_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.gestor_carga_flete_total

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.gestor_carga_flete_total_ml

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_gestor_carga_valor

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_gestor_carga_moneda_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_gestor_carga_unidad_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_gestor_carga_tolerancia

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_gestor_carga_merma

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_gestor_carga_valor_merma

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.total_complemento_a_pagar

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.total_complemento_a_cobrar

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.diferencia_complemento

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.total_descuento_a_pagar

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.total_descuento_a_cobrar

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.diferencia_descuento

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.total_anticipo_retirado

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.saldo_gestor_carga

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.saldo_propietario

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.fecha_conciliacion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.created_by

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.created_at

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.modified_by

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.modified_at

    ws.auto_filter.ref = ws.dimensions
    filename = "rentabilidad_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
