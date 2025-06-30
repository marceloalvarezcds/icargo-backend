import os
from typing import Optional

from fastapi import HTTPException
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories, schemas
from app.config import REPORTS_FOLDER
from app.enums import EstadoEnum
from app.models import Flete

from .flete_anticipo import update_flete_anticipo_list
from .flete_complemento import update_flete_complemento_list
from .flete_descuento import update_flete_descuento_list
from .flete_destinatario import (
    get_destinatario_selected_list_by_flete,
    update_flete_destinatario_list,
)

from sqlalchemy import text



def get_flete_datail_by_id(db: Session, id: int) -> Flete:
    obj = repositories.get_flete_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Flete no encontrado")
    return get_flete_detail(obj)


def get_flete_detail(model: Flete) -> schemas.Flete:
    obj = schemas.Flete.from_orm(model)
    obj.destinatarios = get_destinatario_selected_list_by_flete(model)
    return obj


def reset_sequence(db, seq_name: str, table_name: str):
    max_id = db.execute(text(f"SELECT MAX(id) FROM {table_name}")).scalar()
    if max_id is None:
        db.execute(text(f"SELECT setval('{seq_name}', 1, false)"))
    else:
        db.execute(text(f"SELECT setval('{seq_name}', {max_id}, true)"))


def create_flete(
    db: Session,
    data: schemas.FleteForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> schemas.Flete:
    reset_sequence(db, 'flete_id_seq', 'flete')
    reset_sequence(db, 'flete_anticipo_id_seq', 'flete_anticipo')
    reset_sequence(db, 'flete_complemento_id_seq', 'flete_complemento')
    reset_sequence(db, 'flete_descuento_id_seq', 'flete_descuento')

    obj = repositories.create_flete(
        db,
        data,
        gestor_carga_id,
        modified_by,
    )
    update_flete_anticipo_list(db, data.anticipos, obj, modified_by)
    update_flete_complemento_list(db, data.complementos, obj, modified_by)
    update_flete_descuento_list(db, data.descuentos, obj, modified_by)
    update_flete_destinatario_list(db, data.destinatarios, obj, modified_by)
    return get_flete_detail(obj)



def get_flete_by_id(db: Session, id: int) -> Flete:
    obj = repositories.get_flete_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Flete no encontrado")
    return obj


def get_flete_detail_by_id(db: Session, id: int) -> schemas.FleteList:
    return get_flete_by_id(db, id)


def update_flete_cantidad(
    id: int,
    condicion_cantidad: int,
    db: Session,
    modified_by: str,
) -> schemas.Flete:
    obj = get_flete_by_id(db, id)

    # Calcular diferencia entre nueva y original
    diferencia = condicion_cantidad - obj.condicion_cantidad

    # Actualizar la cantidad
    obj.condicion_cantidad = condicion_cantidad

    # Ajustar el saldo sumando la diferencia
    obj.saldo += diferencia

    obj.modified_by = modified_by
    db.commit()
    db.refresh(obj)
    return get_flete_detail(obj)



def edit_flete(
    id: int,
    db: Session,
    data: schemas.FleteForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> schemas.Flete:
    to_edit_obj = get_flete_by_id(db, id)
    obj = repositories.edit_flete(
        to_edit_obj,
        db,
        data,
        gestor_carga_id,
        modified_by,
    )
    update_flete_anticipo_list(db, data.anticipos, obj, modified_by)
    update_flete_complemento_list(db, data.complementos, obj, modified_by)
    update_flete_descuento_list(db, data.descuentos, obj, modified_by)
    update_flete_destinatario_list(db, data.destinatarios, obj, modified_by)
    return get_flete_detail(obj)


def delete_flete(db: Session, id: int, modified_by: str) -> schemas.Flete:
    co = get_flete_by_id(db, id)
    co.publicado = False
    return get_flete_detail(repositories.delete_flete(co, db, modified_by))


def change_flete_status(
    db: Session, id: int, status: EstadoEnum, modified_by: str
) -> schemas.Flete:
    co = get_flete_by_id(db, id)
    # co.publicado = False
    return get_flete_detail(
        repositories.change_flete_status(co, db, status, modified_by)
    )


def change_flete_public_status(
    db: Session, id: int, is_public: bool, modified_by: str
) -> schemas.Flete:
    co = get_flete_by_id(db, id)
    if is_public and co.estado == EstadoEnum.CANCELADO.value:
        raise HTTPException(
            status_code=403,
            detail="El Flete ya está cancelado, no puede publicarlo",
        )
    return get_flete_detail(
        repositories.change_flete_public_status(co, db, is_public, modified_by)
    )


def get_flete_reports(db: Session) -> str:
    datalist = repositories.get_flete_list(db)
    wb = Workbook()
    # get worksheet
    ws = wb.active
    i = 0

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Nº"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Remitente"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Producto"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tipo de Carga"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Número de Lote"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Publicado"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tipo de Pedido"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Estado"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Gestor de Cuenta"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Origen"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Origen Indicaciones"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Destino"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Destino Indicaciones"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Distancia"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tipo de flete"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Cantidad a Transportar"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Condición para Gestor - Moneda"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Condición para Gestor - Tarifa"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Condición para Gestor - Unidad"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Condición para Propietario - Moneda"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Condición para Propietario - Tarifa"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Condición para Propietario - Unidad"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Merma para Gestor - Valor"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Merma para Gestor - Moneda"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Merma para Gestor - Unidad"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Merma para Gestor - Es Cálculo porcentual"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Merma para Gestor - Tolerancia"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Merma para Propietario - Valor"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Merma para Propietario - Moneda"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Merma para Propietario - Unidad"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Merma para Propietario - Es Cálculo porcentual"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Merma para Propietario - Tolerancia"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Usuario creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Usuario modificación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha modificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        i = 0
        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.id

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.remitente_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.producto_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.tipo_carga_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.numero_lote

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.publicado_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = "Subasta" if item.es_subasta else "Flete"

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.estado

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.gestor_carga_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.origen_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.origen_indicacion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.destino_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.destino_indicacion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.distancia

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.tipo_flete.value

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_cantidad

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_gestor_carga_moneda_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_gestor_cuenta_tarifa

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_gestor_carga_unidad_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_propietario_moneda_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_propietario_tarifa

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.condicion_propietario_unidad_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_gestor_cuenta_valor

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_gestor_carga_moneda_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_gestor_carga_unidad_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = "Si" if item.merma_gestor_cuenta_es_porcentual else "No"

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_gestor_cuenta_tolerancia

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_propietario_valor

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_propietario_moneda_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_propietario_unidad_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = "Si" if item.merma_propietario_es_porcentual else "No"

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.merma_propietario_tolerancia

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.created_by

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.created_at

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.modified_by

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.modified_at

    ws.auto_filter.ref = ws.dimensions
    filename = "flete_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
