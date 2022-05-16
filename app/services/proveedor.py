import os
from typing import List, Optional

from fastapi import HTTPException, UploadFile  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories, schemas
from app.config import REPORTS_FOLDER
from app.models import GestorCargaProveedor, Proveedor, ProveedorContactoGestorCarga

from .gestor_carga_proveedor import (
    create_gestor_carga_proveedor,
    edit_gestor_carga_proveedor,
)
from .pictshare import upload_and_get_image_url
from .proveedor_contacto import update_proveedor_contacto_list


def get_proveedor_detail(
    obj: Proveedor, gestor_carga_id: Optional[int]
) -> schemas.Proveedor:
    obj_dict = obj.as_dict(for_json=False)
    contactos: List[ProveedorContactoGestorCarga] = obj.contactos
    ges: List[GestorCargaProveedor] = obj.gestores
    gestores = [x for x in ges if x.gestor_carga_id == gestor_carga_id]
    obj_dict["contactos"] = [
        x for x in contactos if x.gestor_carga_id == gestor_carga_id
    ]
    obj_dict["gestor_carga_proveedor"] = gestores[0] if len(gestores) > 0 else None
    obj_dict["tipo_documento"] = obj.tipo_documento
    obj_dict["composicion_juridica"] = obj.composicion_juridica
    obj_dict["ciudad"] = obj.ciudad
    return schemas.Proveedor.parse_obj(obj_dict)


async def create_proveedor(
    db: Session,
    data: schemas.ProveedorForm,
    file: Optional[UploadFile],
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> schemas.Proveedor:
    if repositories.get_proveedor_by(db, data.tipo_documento_id, data.numero_documento):
        raise HTTPException(
            status_code=409,
            detail=f"El Proveedor con documento {data.numero_documento} ya existe",
        )
    logo_url = await upload_and_get_image_url(file)
    obj = repositories.create_proveedor(db, data, logo_url, modified_by)
    update_proveedor_contacto_list(
        db, data.contactos, obj, gestor_carga_id, modified_by
    )
    create_gestor_carga_proveedor(db, obj, gestor_carga_id, data.alias, modified_by)
    return get_proveedor_detail(obj, gestor_carga_id)


def get_proveedor_by_id(db: Session, id: int) -> Proveedor:
    obj = repositories.get_proveedor_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    return obj


def get_proveedor_by_id_and_gestor_carga_id(
    db: Session, id: int, gestor_carga_id: Optional[int] = None
) -> schemas.Proveedor:
    obj = repositories.get_proveedor_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    return get_proveedor_detail(obj, gestor_carga_id)


async def edit_proveedor(
    id: int,
    db: Session,
    data: schemas.ProveedorForm,
    file: Optional[UploadFile],
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> schemas.Proveedor:
    exists = repositories.get_proveedor_by(
        db, data.tipo_documento_id, data.numero_documento
    )
    if exists and exists.id != id:
        raise HTTPException(
            status_code=409,
            detail=f"El Proveedor con documento {data.numero_documento} ya existe",
        )
    logo_url = await upload_and_get_image_url(file) if file else None
    to_edit_obj = get_proveedor_by_id(db, id)
    obj = repositories.edit_proveedor(to_edit_obj, db, data, logo_url, modified_by)
    update_proveedor_contacto_list(
        db, data.contactos, obj, gestor_carga_id, modified_by
    )
    edit_gestor_carga_proveedor(db, obj, gestor_carga_id, data.alias, modified_by)
    return get_proveedor_detail(obj, gestor_carga_id)


def delete_proveedor(
    db: Session, id: int, gestor_carga_id: Optional[int], modified_by: str
) -> schemas.Proveedor:
    co = get_proveedor_by_id(db, id)
    obj = repositories.delete_proveedor(co, db, modified_by)
    return get_proveedor_detail(obj, gestor_carga_id)


def get_proveedor_reports(db: Session) -> str:
    datalist = repositories.get_proveedor_list(db)
    wb = Workbook()
    # get worksheet
    ws = wb.active

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Nombre"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Nombre de Fantasía"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "Tipo de Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Número de Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Composición Jurídica"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=6)
    title_cell.value = "Dirección"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Ubicación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.nombre

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.nombre_corto if item.nombre_corto else ""

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.tipo_documento.descripcion

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.numero_documento

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = (
            item.composicion_juridica.nombre if item.composicion_juridica else ""
        )

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.direccion if item.direccion else ""

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = (
            f"{item.ciudad.nombre}/{item.ciudad.localidad.nombre}/{item.ciudad.localidad.pais.nombre_corto}"  # noqa: B950
            if item.ciudad
            else ""
        )

    ws.auto_filter.ref = ws.dimensions
    filename = "proveedor_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
