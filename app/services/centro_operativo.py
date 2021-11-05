import os
from typing import List, Optional

from fastapi import HTTPException, UploadFile  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories, schemas
from app.config import REPORTS_FOLDER
from app.models import (
    CentroOperativo,
    CentroOperativoContactoGestorCarga,
    GestorCargaCentroOperativo,
)

from .centro_operativo_contacto import update_centro_operativo_contacto_list
from .gestor_carga_centro_operativo import (
    create_gestor_carga_centro_operativo,
    edit_gestor_carga_centro_operativo,
)
from .pictshare import upload_and_get_image_url


def get_centro_operativo_detail(
    obj: CentroOperativo, gestor_carga_id: Optional[int]
) -> schemas.CentroOperativo:
    obj_dict = obj.as_dict(for_json=False)
    contactos: List[CentroOperativoContactoGestorCarga] = obj.contactos
    ges: List[GestorCargaCentroOperativo] = obj.gestores
    gestores = [x for x in ges if x.gestor_carga_id == gestor_carga_id]
    obj_dict["contactos"] = [
        x for x in contactos if x.gestor_carga_id == gestor_carga_id
    ]
    obj_dict["gestor_carga_centro_operativo"] = (
        gestores[0] if len(gestores) > 0 else None
    )
    obj_dict["clasificacion"] = obj.clasificacion
    obj_dict["ciudad"] = obj.ciudad
    return schemas.CentroOperativo.parse_obj(obj_dict)


async def create_centro_operativo(
    db: Session,
    data: schemas.CentroOperativoForm,
    file: UploadFile,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> schemas.CentroOperativo:
    logo_url = await upload_and_get_image_url(file)
    obj = repositories.create_centro_operativo(db, data, logo_url, modified_by)
    update_centro_operativo_contacto_list(
        db, data.contactos, obj, gestor_carga_id, modified_by
    )
    create_gestor_carga_centro_operativo(
        db, obj, gestor_carga_id, data.alias, modified_by
    )
    return get_centro_operativo_detail(obj, gestor_carga_id)


def get_centro_operativo_by_id(db: Session, id: int) -> CentroOperativo:
    obj = repositories.get_centro_operativo_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Centro Operativo no encontrado")
    return obj


def get_centro_operativo_by_id_and_gestor_carga_id(
    db: Session, id: int, gestor_carga_id: Optional[int] = None
) -> schemas.CentroOperativo:
    obj = repositories.get_centro_operativo_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Centro Operativo no encontrado")
    return get_centro_operativo_detail(obj, gestor_carga_id)


async def edit_centro_operativo(
    id: int,
    db: Session,
    data: schemas.CentroOperativoForm,
    file: Optional[UploadFile],
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> schemas.CentroOperativo:
    logo_url = await upload_and_get_image_url(file) if file else None
    to_edit_obj = get_centro_operativo_by_id(db, id)
    obj = repositories.edit_centro_operativo(
        to_edit_obj, db, data, logo_url, modified_by
    )
    update_centro_operativo_contacto_list(
        db, data.contactos, obj, gestor_carga_id, modified_by
    )
    edit_gestor_carga_centro_operativo(
        db, obj, gestor_carga_id, data.alias, modified_by
    )
    return get_centro_operativo_detail(obj, gestor_carga_id)


def delete_centro_operativo(
    db: Session, id: int, gestor_carga_id: Optional[int], modified_by: str
) -> schemas.CentroOperativo:
    co = get_centro_operativo_by_id(db, id)
    obj = repositories.delete_centro_operativo(co, db, modified_by)
    return get_centro_operativo_detail(obj, gestor_carga_id)


def get_centro_operativo_reports(db: Session) -> str:
    datalist = repositories.get_centro_operativo_list(db)
    wb = Workbook()
    # get worksheet
    ws = wb.active
    # title = f"Lista de Centros Operativos"
    # title_cell = ws.cell(row=1, column=1)
    # title_cell.font = Font(size=12, bold=True)
    # title_cell.value = title.upper()

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Nombre"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Nombre de Fantasía"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "Dirección"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Ubicación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Clasificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.nombre

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.nombre_corto if item.nombre_corto else ""

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.direccion if item.direccion else ""

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = f"{item.ciudad.nombre}/{item.ciudad.localidad.nombre}/{item.ciudad.localidad.pais.nombre_corto}"  # noqa

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = item.clasificacion.nombre

    ws.auto_filter.ref = ws.dimensions
    filename = "centro_operativo_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
