import os
from datetime import datetime
from http import HTTPStatus
from typing import List, Optional, cast

from app.models.combinacion import Combinacion
from fastapi import HTTPException
from jinja2 import Template
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from pdfkit import from_string  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories, schemas
from app.config import LOGO_IMAGE_URL, REPORTS_FOLDER, STATICS_URL, templateEnv
from app.enums import EstadoEnum
from app.models import Camion, Flete, GestorCarga, OrdenCarga
from app.schemas.audit_database import AuditDatabase as A
from app.utils import number_format, send_email_with_template_by_thread

from .audit_database import get_audit_list_by_orden_carga
from .camion_semi_neto import (
    get_camion_semi_neto_by_camion_id_and_semi_id_and_producto_id,
)
from .flete import get_flete_detail
from .movimiento import create_movimiento_by_conciliacion_oc
from .orden_carga_anticipo_porcentaje import (
    create_orden_carga_anticipo_porcentaje_by_flete_anticipo_list,
    edit_orden_carga_anticipo_porcentaje_by_oc_porcentaje_anticipos,
)
from .orden_carga_anticipo_saldo import get_orden_carga_by_id, get_orden_carga_by_combinacion_id
from .orden_carga_complemento_flete import create_orden_carga_complemento_by_flete
from .orden_carga_descuento_flete import create_orden_carga_descuento_by_flete
from .orden_carga_remision_resultado import (
    get_orden_carga_remision_resultado_list_by_flete,
    get_orden_carga_remision_resultado_list_by_orden_carga,
)


def get_orden_carga_list(
    db: Session, gestor_carga_id: Optional[int]
) -> List[OrdenCarga]:
    if gestor_carga_id:
        return repositories.get_orden_carga_list_by_gestor_carga_id(db, gestor_carga_id)
    return repositories.get_orden_carga_list(db)

def get_orden_carga_list_combinacion(
    db: Session, combinacion_id, gestor_carga_id: Optional[int]
) -> List[OrdenCarga]:
    if gestor_carga_id:
        return repositories.get_orden_de_carga_by_combinacion_id(db, combinacion_id, gestor_carga_id)
    return repositories.get_orden_carga_list(db)


def get_orden_carga_with_resultado(
    db: Session, model: OrdenCarga, user_id: int
) -> schemas.OrdenCarga:
    obj = schemas.OrdenCarga.from_orm(model)
    obj.remisiones_resultado = get_orden_carga_remision_resultado_list_by_orden_carga(
        db, model, user_id
    )
    obj.remisiones_resultado_flete = get_orden_carga_remision_resultado_list_by_flete(
        db, model, user_id
    )
    obj.auditorias = cast(List[A], get_audit_list_by_orden_carga(db, model, user_id))
    return obj


def create_complementos_and_descuentos(
    db: Session, obj: OrdenCarga, flete: Flete, modified_by: str
):
    for c in flete.complementos:
        create_orden_carga_complemento_by_flete(db, obj, c, modified_by)
    for d in flete.descuentos:
        create_orden_carga_descuento_by_flete(db, obj, d, modified_by)
        

def create_orden_carga(
    db: Session,
    data: schemas.OrdenCargaForm,
    current_user: schemas.AuthUser,
) -> schemas.OrdenCarga:
    flete = repositories.get_flete_by_id(db, data.flete_id)
    if not flete:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
    gestor_carga_id = current_user.gestor_carga_id
    modified_by = current_user.username
    camion_semi_neto = get_camion_semi_neto_by_camion_id_and_semi_id_and_producto_id(
        db, data.camion_id, data.semi_id, flete.producto_id, gestor_carga_id
    )
    data.camion_semi_neto_id = camion_semi_neto.id if camion_semi_neto else None
    obj = repositories.create_orden_carga(
        db,
        data,
        flete,
        gestor_carga_id,
        modified_by,
    )
    create_orden_carga_anticipo_porcentaje_by_flete_anticipo_list(
        db, obj.id, obj.flete_anticipos, modified_by
    )
    create_complementos_and_descuentos(db, obj, flete, modified_by)
    return get_orden_carga_with_resultado(db, obj, current_user.id)


def edit_orden_carga(
    id: int,
    db: Session,
    data: schemas.OrdenCargaEditForm,
    current_user: schemas.AuthUser,
) -> schemas.OrdenCarga:
    to_edit_obj = get_orden_carga_by_id(db, id)
    obj = repositories.edit_orden_carga(
        to_edit_obj,
        db,
        data,
        current_user.gestor_carga_id,
        current_user.username,
    )
    edit_orden_carga_anticipo_porcentaje_by_oc_porcentaje_anticipos(
        obj.id,
        db,
        to_edit_obj.porcentaje_anticipos,
        data.porcentaje_anticipos,
        current_user.username,
    )
    return get_orden_carga_with_resultado(db, obj, current_user.id)


def update_comentarios(
    id: int,
    db: Session,
    data: schemas.OrdenCargaUpdateForm,
    current_user: schemas.AuthUser,
) -> schemas.OrdenCarga:
    orden_carga = db.query(OrdenCarga).filter(OrdenCarga.id == id).first()
    if not orden_carga:
        raise HTTPException(status_code=404, detail="Orden de carga no encontrada")

    if data.comentarios is not None:
        orden_carga.comentarios = data.comentarios

    db.commit()
    db.refresh(orden_carga)

    return orden_carga


def delete_orden_carga(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    model = repositories.delete_orden_carga(obj, db, current_user.username)
    return get_orden_carga_with_resultado(db, model, current_user.id)


def get_orden_carga_detail(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    return get_orden_carga_with_resultado(db, obj, current_user.id)


def get_orden_carga_combinacion_detail(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj =  get_orden_carga_by_combinacion_id(db, id)
    return get_orden_carga_with_resultado(db, obj, current_user.id)


def get_ordenes_carga_by_combinacion_id(db: Session, combinacion_id: int) -> List[OrdenCarga]:
    ordenes_carga = db.query(OrdenCarga).filter(OrdenCarga.combinacion_id == combinacion_id).all()
    if not combinacion_id:
        raise HTTPException(status_code=404, detail="No se encontraron órdenes de carga para esta combinación")
    return ordenes_carga


def get_ordenes_carga_by_combinacion_id_and_nuevo(db: Session, combinacion_id: int) -> List[OrdenCarga]:
    if not combinacion_id:
        raise HTTPException(status_code=404, detail="No se encontró la combinación ID proporcionada")
    
    # Filtrar por combinacion_id y estado "Aceptado"
    ordenes_carga = db.query(OrdenCarga).filter(
        OrdenCarga.combinacion_id == combinacion_id,
        OrdenCarga.estado == "Nuevo"  # Ajusta este valor si el estado se almacena de manera diferente
    ).all()
    
    return ordenes_carga


def get_ordenes_carga_by_combinacion_id_and_finalizar(db: Session, combinacion_id: int) -> List[OrdenCarga]:
    if not combinacion_id:
        raise HTTPException(status_code=404, detail="No se encontró la combinación ID proporcionada")
    
    # Filtrar por combinacion_id y estado "Aceptado"
    ordenes_carga = db.query(OrdenCarga).filter(
        OrdenCarga.combinacion_id == combinacion_id,
        OrdenCarga.estado == "Finalizado"  # Ajusta este valor si el estado se almacena de manera diferente
    ).all()
    
    return ordenes_carga


def get_ordenes_carga_by_combinacion_id_and_aceptado(db: Session, combinacion_id: int) -> List[OrdenCarga]:
    if not combinacion_id:
        raise HTTPException(status_code=404, detail="No se encontró la combinación ID proporcionada")
    
    # Filtrar por combinacion_id y estado "Aceptado"
    ordenes_carga = db.query(OrdenCarga).filter(
        OrdenCarga.combinacion_id == combinacion_id,
        OrdenCarga.estado == "Aceptado"  # Ajusta este valor si el estado se almacena de manera diferente
    ).all()
    
    return ordenes_carga


def get_orden_carga_list_by_combinacion_id(
    db: Session, id: int
) -> List[OrdenCarga]:
    obj = repositories.get_orden_carga_by_combinacion_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Combinacion no encontrada")
    return obj



def get_orden_carga_pdf_by_id(db: Session, id: int) -> str:
    obj = repositories.get_orden_carga_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Orden de Carga no encontrada")
    gestor_carga = repositories.get_gestor_carga_by_id(db, obj.gestor_carga_id)
    if not gestor_carga:
        raise HTTPException(status_code=404, detail="Gestor no encontrado")
    OUTPUT_FILENAME = f"orden_carga_{id}.pdf"
    TEMPLATE_FILENAME = "pdf_orden_carga.html"
    template: Template = templateEnv.get_template(TEMPLATE_FILENAME)
    create_at: datetime = obj.created_at
    fecha_vencimiento: datetime = obj.camion.vencimiento_habilitacion_transporte
    df = "%Y-%m-%d / %H:%M:%S"
    bruto = obj.camion.bruto
    neto = obj.camion_semi_neto.neto if obj.camion_semi_neto else None
    gestor_carga_nombre_corto = (
        gestor_carga.nombre_corto if gestor_carga.nombre_corto else gestor_carga.nombre
    )
    data = {
        "id": id,
        "flete_id": obj.flete_id,
        "remitente": obj.flete.remitente.nombre,
        "remitente_numero_documento": obj.flete.remitente.numero_documento,
        "bruto": number_format(bruto) if bruto else "-",
        "neto": number_format(neto) if neto else "-",
        "cantidad_nominada": number_format(obj.cantidad_nominada),
        "gestor_carga_direccion": gestor_carga.direccion,
        "gestor_carga_logo": gestor_carga.logo,
        "gestor_carga_nombre": gestor_carga.nombre,
        "gestor_carga_nombre_corto": gestor_carga_nombre_corto,
        "gestor_carga_numero_documento": gestor_carga.numero_documento,
        "fecha": datetime.now().strftime(df),
        "fecha_orden": create_at.strftime(df),
        "fecha_validez": obj.fecha_validez.strftime(df),
        "producto": obj.flete_producto_descripcion,
        "origen": obj.origen_nombre,
        "origen_direccion": obj.origen.direccion if obj.origen.direccion else "-",
        "destino": obj.destino_nombre,
        "destino_direccion": obj.destino.direccion if obj.destino.direccion else "-",
        "propietario_nombre": obj.camion_propietario_nombre,
        "propietario_telefono": obj.camion.propietario.telefono,
        "chofer_nombre": obj.camion_chofer_nombre,
        "chofer_numero_documento": obj.camion_chofer_numero_documento,
        "chofer_telefono": obj.camion.chofer.telefono,
        "camion_foto": obj.camion.foto,
        "camion_placa": obj.camion_placa,
        "camion_marca_tipo": f"{obj.camion.marca_descripcion}/{obj.camion.tipo_descripcion}",
        "camion_color": obj.camion.color_descripcion,
        "camion_tipo": obj.camion.tipo_descripcion,
        "camion_fecha_vto": fecha_vencimiento.strftime(df)
        if fecha_vencimiento
        else "-",
        "semi_placa": obj.semi_placa,
        "semi_marca_tipo": f"{obj.semi.marca_descripcion}/{obj.semi.tipo_descripcion}",
        "semi_color": obj.semi.color_descripcion,
        "semi_tipo": obj.semi.tipo_descripcion,
        "comentarios": obj.comentarios if obj.comentarios else "-",
        "texto_legal": obj.flete.emision_orden_texto_legal
        if obj.flete.emision_orden_texto_legal
        else "-",
        "usuario": obj.created_by,
    }
    source_html = template.render(logo=LOGO_IMAGE_URL, times=range(2), **data)
    pdf_filename = os.path.join(REPORTS_FOLDER, OUTPUT_FILENAME)
    from_string(source_html, pdf_filename, {"page-size": "Legal"})
    return OUTPUT_FILENAME


def get_orden_carga_resumen_pdf_by_id(db: Session, id: int) -> str:
    obj = repositories.get_orden_carga_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Orden de Carga no encontrada")
    gestor_carga = repositories.get_gestor_carga_by_id(db, obj.gestor_carga_id)
    if not gestor_carga:
        raise HTTPException(status_code=404, detail="Gestor no encontrado")
    marca_agua_url = f"{STATICS_URL}/marca-de-agua.png"
    OUTPUT_FILENAME = f"resumen_{id}.pdf"
    TEMPLATE_FILENAME = "pdf_resumen.html"
    template: Template = templateEnv.get_template(TEMPLATE_FILENAME)
    data = {
        "id": id,
        "flete_id": obj.flete_id,
        "gestor_carga_direccion": gestor_carga.direccion,
        "gestor_carga_logo": gestor_carga.logo,
        "gestor_carga_nombre": gestor_carga.nombre,
        "gestor_carga_numero_documento": gestor_carga.numero_documento,
        "fecha": datetime.now().strftime("%Y-%m-%d / %H:%M:%S"),
        "propietario_nombre": obj.camion_propietario_nombre,
        "chofer_nombre": obj.camion_chofer_nombre,
        "camion_placa": obj.camion_placa,
        "semi_placa": obj.semi_placa,
        "origen": obj.origen_nombre,
        "origen_direccion": obj.origen.direccion if obj.origen.direccion else "-",
        "destino": obj.destino_nombre,
        "destino_direccion": obj.destino.direccion if obj.destino.direccion else "-",
        "producto": obj.flete_producto_descripcion,
        "tarifa_flete": number_format(obj.flete_tarifa),
        "tasa": obj.flete_tarifa_unidad,  # noqa
        "docs_origen": obj.remisiones,
        "cantidad_origen": number_format(obj.cantidad_origen),
        "docs_destino": obj.nro_tickets,
        "cantidad_destino": number_format(obj.cantidad_destino),
        "diferencia": number_format(obj.diferencia_origen_destino),
        "tolerancia": number_format(obj.resultado_propietario_tolerancia_kg),
        "merma": number_format(obj.resultado_propietario_merma),
        "total_flete": number_format(obj.flete_proyectado),
        "complementos": number_format(obj.resultado_propietario_total_complemento),
        "descuentos": number_format(obj.resultado_propietario_total_descuento),
        "merma_total": number_format(
            obj.resultado_propietario_merma_valor_total_moneda_local
        ),
        "anticipo": number_format(obj.resultado_propietario_total_anticipos_retirados),
        "saldo": number_format(obj.resultado_propietario_saldo),
        "marca_agua_url": marca_agua_url,
        "class_name": "marca-agua" if obj.estado == EstadoEnum.FINALIZADO.value else "",
    }
    source_html = template.render(logo=LOGO_IMAGE_URL, **data)
    pdf_filename = os.path.join(REPORTS_FOLDER, OUTPUT_FILENAME)
    from_string(
        source_html, pdf_filename, {"page-size": "Legal", "orientation": "Landscape"}
    )
    return OUTPUT_FILENAME


def change_orden_carga_anticipos_liberados(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    if not obj.anticipos_liberados and obj.estado == EstadoEnum.CANCELADO.value:
        raise HTTPException(
            status_code=403,
            detail="La orden de carga ya está cancelada, no puede liberar anticipos",
        )
    anticipos_liberados = not obj.anticipos_liberados
    model = repositories.change_orden_carga_anticipos_liberados(
        obj, db, anticipos_liberados, current_user.username
    )
    return get_orden_carga_with_resultado(db, model, current_user.id)


def aceptar_orden_carga(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    cant_oc_aceptadas = repositories.get_orden_carga_aceptada_count_by_camion_id(
        db, obj.camion_id
    )
    camion: Camion = obj.camion
    camion.limite_cantidad_oc_activas
    if cant_oc_aceptadas >= camion.limite_cantidad_oc_activas:
        oc_aceptadas = f"Existen {cant_oc_aceptadas} OC aceptadas"
        limite = f"y el límite es de {camion.limite_cantidad_oc_activas}"
        camion_placa = f"establecido por Camión con placa {camion.placa}"
        raise HTTPException(
            status_code=HTTPStatus.LOCKED,
            detail=f"{oc_aceptadas} {limite}, {camion_placa}",
        )
    username = current_user.username
    gestor_id = (
        current_user.gestor_carga_id
        if current_user.gestor_carga_id
        else obj.gestor_carga_id
    )
    anticipos_liberados = (
        camion.propietario_puede_recibir_anticipos
        and camion.chofer_puede_recibir_anticipos
    )
    oc = repositories.edit_orden_carga(
        obj,
        db,
        schemas.OrdenCargaEditForm(anticipos_liberados=anticipos_liberados),
        gestor_id,
        username,
    )
    model = repositories.aceptar_orden_carga(oc, db, username)
    send_emision_orden_carga_mail(obj)
    return get_orden_carga_with_resultado(db, model, current_user.id)


def cancelar_orden_carga(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    model = repositories.cancelar_orden_carga(obj, db, current_user.username)
    return get_orden_carga_with_resultado(db, model, current_user.id)


def conciliar_orden_carga(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    model = repositories.conciliar_orden_carga(obj, db, current_user.username)
    create_movimiento_by_conciliacion_oc(
        db, obj, current_user.gestor_carga_id, current_user.username
    )
    return get_orden_carga_with_resultado(db, model, current_user.id)


def contabilizar_orden_carga(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    model = repositories.contabilizar_orden_carga(obj, db, current_user.username)
    return get_orden_carga_with_resultado(db, model, current_user.id)


def arribado_a_cargar_orden_carga(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    model = repositories.arribado_a_cargar_orden_carga(obj, db, current_user.username)
    return get_orden_carga_with_resultado(db, model, current_user.id)


def arribado_a_descargar_orden_carga(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    model = repositories.arribado_a_descargar_orden_carga(
        obj, db, current_user.username
    )
    return get_orden_carga_with_resultado(db, model, current_user.id)


def cargar_orden_carga(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    model = repositories.cargar_orden_carga(obj, db, current_user.username)
    return get_orden_carga_with_resultado(db, model, current_user.id)


def descargar_orden_carga(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    model = repositories.descargar_orden_carga(obj, db, current_user.username)
    return get_orden_carga_with_resultado(db, model, current_user.id)


def finalizar_orden_carga(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    model = repositories.finalizar_orden_carga(obj, db, current_user.username)
    return get_orden_carga_with_resultado(db, model, current_user.id)


def liquidar_orden_carga(
    db: Session, id: int, current_user: schemas.AuthUser
) -> schemas.OrdenCarga:
    obj = get_orden_carga_by_id(db, id)
    model = repositories.liquidar_orden_carga(obj, db, current_user.username)
    return get_orden_carga_with_resultado(db, model, current_user.id)


def send_oc_mail(db: Session, id: int):
    obj = get_orden_carga_by_id(db, id)
    send_emision_orden_carga_mail(obj)


def send_emision_orden_carga_mail(obj: OrdenCarga):
    gestor_carga: GestorCarga = obj.gestor_carga
    flete = get_flete_detail(obj.flete)
    destinatarios = ",".join([x.email for x in flete.destinatarios])
    create_at: datetime = obj.created_at
    fecha_vencimiento: datetime = obj.camion.vencimiento_habilitacion_transporte
    df = "%Y-%m-%d / %H:%M:%S"
    camion_bruto = obj.camion.bruto if obj.camion.bruto else 0
    camion_neto = obj.camion.neto if obj.camion.neto else 0
    semi_bruto = obj.semi.bruto if obj.semi.bruto else 0
    semi_neto = obj.semi.neto if obj.semi.neto else 0
    bruto = camion_bruto + semi_bruto
    neto = camion_neto + semi_neto
    gestor_carga_nombre_corto = (
        gestor_carga.nombre_corto if gestor_carga.nombre_corto else gestor_carga.nombre
    )
    data = {
        "id": obj.id,
        "flete_id": flete.id,
        "remitente": flete.remitente.nombre,
        "remitente_numero_documento": flete.remitente.numero_documento,
        "bruto": number_format(bruto),
        "neto": number_format(neto),
        "cantidad_nominada": number_format(obj.cantidad_nominada),
        "gestor_carga_direccion": gestor_carga.direccion,
        "gestor_carga_logo": gestor_carga.logo,
        "gestor_carga_nombre": gestor_carga.nombre,
        "gestor_carga_nombre_corto": gestor_carga_nombre_corto,
        "gestor_carga_numero_documento": gestor_carga.numero_documento,
        "fecha": datetime.now().strftime(df),
        "fecha_orden": create_at.strftime(df),
        "fecha_validez": obj.fecha_validez.strftime(df),
        "producto": obj.flete_producto_descripcion,
        "origen": obj.origen_nombre,
        "origen_direccion": obj.origen.direccion if obj.origen.direccion else "-",
        "destino": obj.destino_nombre,
        "destino_direccion": obj.destino.direccion if obj.destino.direccion else "-",
        "propietario_nombre": obj.camion_propietario_nombre,
        "propietario_telefono": obj.camion.propietario.telefono,
        "chofer_nombre": obj.camion_chofer_nombre,
        "chofer_numero_documento": obj.camion_chofer_numero_documento,
        "chofer_telefono": obj.camion.chofer.telefono if obj.camion.chofer else "",
        "camion_foto": obj.camion.foto,
        "camion_placa": obj.camion_placa,
        "camion_marca_tipo": f"{obj.camion.marca_descripcion}/{obj.camion.tipo_descripcion}",
        "camion_color": obj.camion.color_descripcion,
        "camion_tipo": obj.camion.tipo_descripcion,
        "camion_fecha_vto": fecha_vencimiento.strftime(df)
        if fecha_vencimiento
        else "-",
        "semi_placa": obj.semi_placa,
        "semi_marca_tipo": f"{obj.semi.marca_descripcion}/{obj.semi.tipo_descripcion}",
        "semi_color": obj.semi.color_descripcion,
        "semi_tipo": obj.semi.tipo_descripcion,
        "comentarios": obj.comentarios if obj.comentarios else "-",
        "texto_legal": obj.flete.emision_orden_texto_legal
        if obj.flete.emision_orden_texto_legal
        else "-",
        "usuario": obj.created_by,
    }
    send_email_with_template_by_thread(
        template_filename="mail_orden_carga.html",
        to=destinatarios,
        subject="iCargo: Emisión de Orden de Carga",
        logo=LOGO_IMAGE_URL,
        **data,
    )


def get_orden_carga_reports(db: Session, gestor_carga_id: Optional[int]) -> str:
    datalist = repositories.get_orden_carga_list_by_gestor_carga_id(db, gestor_carga_id)
    wb = Workbook()
    # get worksheet
    ws = wb.active

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Nº"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Estado"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "Estado Anticipos"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Nº Pedido"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Placa Camión"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=6)
    title_cell.value = "Placa Semirrelque"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Cant. Nominada (kg)"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=8)
    title_cell.value = "Comentarios"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=9)
    title_cell.value = "Gestora de Carga"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=10)
    title_cell.value = "Remisiones"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=11)
    title_cell.value = "Nº Ticket"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=12)
    title_cell.value = "Remitente"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=13)
    title_cell.value = "Chofer"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=14)
    title_cell.value = "Propietario"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=15)
    title_cell.value = "Tipo de Flete"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=16)
    title_cell.value = "Producto"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=17)
    title_cell.value = "Origen"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=18)
    title_cell.value = "Destino"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=19)
    title_cell.value = "Cant. Origen (kg)"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=20)
    title_cell.value = "Cant. Destino (kg)"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=21)
    title_cell.value = "Lugar de Carga"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=22)
    title_cell.value = "Lugar de Descarga"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=23)
    title_cell.value = "Usuario creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=24)
    title_cell.value = "Fecha creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=25)
    title_cell.value = "Usuario modificación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=26)
    title_cell.value = "Fecha modificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.id

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.estado

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.anticipos_liberados_descripcion

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.flete_id

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = item.camion_placa

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.semi_placa

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = item.cantidad_nominada

        value_cell = ws.cell(row=row + 2, column=8)
        value_cell.value = item.comentarios

        value_cell = ws.cell(row=row + 2, column=9)
        value_cell.value = item.gestor_carga_nombre

        value_cell = ws.cell(row=row + 2, column=10)
        value_cell.value = item.remisiones

        value_cell = ws.cell(row=row + 2, column=11)
        value_cell.value = item.nro_tickets

        value_cell = ws.cell(row=row + 2, column=12)
        value_cell.value = item.flete_remitente_nombre

        value_cell = ws.cell(row=row + 2, column=13)
        value_cell.value = item.camion_chofer_nombre

        value_cell = ws.cell(row=row + 2, column=14)
        value_cell.value = item.camion_propietario_nombre

        value_cell = ws.cell(row=row + 2, column=15)
        value_cell.value = item.flete_tipo.value if item.flete_tipo else ""

        value_cell = ws.cell(row=row + 2, column=16)
        value_cell.value = item.flete_producto_descripcion

        value_cell = ws.cell(row=row + 2, column=17)
        value_cell.value = item.flete_origen_nombre

        value_cell = ws.cell(row=row + 2, column=18)
        value_cell.value = item.flete_destino_nombre

        value_cell = ws.cell(row=row + 2, column=19)
        value_cell.value = item.cantidad_origen

        value_cell = ws.cell(row=row + 2, column=20)
        value_cell.value = item.cantidad_destino

        value_cell = ws.cell(row=row + 2, column=21)
        value_cell.value = item.origen_nombre

        value_cell = ws.cell(row=row + 2, column=22)
        value_cell.value = item.destino_nombre

        value_cell = ws.cell(row=row + 2, column=23)
        value_cell.value = item.created_by

        value_cell = ws.cell(row=row + 2, column=24)
        value_cell.value = item.created_at

        value_cell = ws.cell(row=row + 2, column=25)
        value_cell.value = item.modified_by

        value_cell = ws.cell(row=row + 2, column=26)
        value_cell.value = item.modified_at

    ws.auto_filter.ref = ws.dimensions
    filename = "orden_carga_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
