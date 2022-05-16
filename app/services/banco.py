import os
from typing import List, Optional

from fastapi import HTTPException  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories
from app.config import REPORTS_FOLDER
from app.models import Banco
from app.schemas import BancoForm


def get_banco_list(db: Session, gestor_carga_id: Optional[int]) -> List[Banco]:
    if gestor_carga_id:
        return repositories.get_banco_list_by_gestor_carga_id(db, gestor_carga_id)
    return repositories.get_banco_list(db)


def create_banco(
    db: Session,
    data: BancoForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Banco:
    gestor_id = gestor_carga_id if gestor_carga_id else data.gestor_carga_id
    if not gestor_id:
        raise HTTPException(status_code=409, detail="Debe elegir un Gestor de carga")
    if repositories.get_banco_by(db, data.numero_cuenta, gestor_id):
        raise HTTPException(
            status_code=409, detail=f"El Banco {data.numero_cuenta} ya existe"
        )
    return repositories.create_banco(db, data, gestor_id, modified_by)


def get_banco_by_id(db: Session, id: int) -> Banco:
    obj = repositories.get_banco_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Banco no encontrado")
    return obj


def edit_banco(
    id: int,
    db: Session,
    data: BancoForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Banco:
    gestor_id = gestor_carga_id if gestor_carga_id else data.gestor_carga_id
    if not gestor_id:
        raise HTTPException(status_code=409, detail="Debe elegir un Gestor de carga")
    exists = repositories.get_banco_by(db, data.numero_cuenta, gestor_id)
    if exists and exists.id != id:
        raise HTTPException(
            status_code=409, detail=f"El Banco {data.numero_cuenta} ya existe"
        )
    to_edit_obj = get_banco_by_id(db, id)
    return repositories.edit_banco(to_edit_obj, db, data, gestor_id, modified_by)


def delete_banco(db: Session, id: int, modified_by: str) -> Banco:
    co = get_banco_by_id(db, id)
    return repositories.delete_banco(co, db, modified_by)


def get_banco_reports(db: Session) -> str:
    datalist = repositories.get_banco_list(db)
    wb = Workbook()
    ws = wb.active
    i = 0

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "ID"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Nombre de Cuenta"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Titular"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Banco"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Saldo"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Pendiente"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Usuario creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Usuario modificación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha modificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        i = 0

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.id

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.numero_cuenta

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.titular

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.saldo_confirmado

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.saldo_provisional

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.created_by

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.created_at

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.modified_by

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.modified_at

    ws.auto_filter.ref = ws.dimensions
    filename = "banco_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
