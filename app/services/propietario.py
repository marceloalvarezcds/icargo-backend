import os
from typing import Optional, cast

from fastapi import HTTPException, UploadFile  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories, schemas
from app.config import REPORTS_FOLDER
from app.enums import EstadoEnum
from app.models import Propietario

from .gestor_carga_propietario import (
    create_gestor_carga_propietario,
    edit_gestor_carga_propietario,
)
from .propietario_check_files import check_files, get_propietario_detail
from .propietario_chofer import (
    create_or_edit_chofer_by_propietario,
    disable_chofer_by_id,
)
from .propietario_contacto import update_propietario_contacto_list


async def create_propietario(
    db: Session,
    data: schemas.PropietarioForm,
    foto_documento_frente_file: UploadFile,
    foto_documento_reverso_file: UploadFile,
    foto_perfil_file: UploadFile,
    foto_documento_frente_chofer_file: Optional[UploadFile],
    foto_documento_reverso_chofer_file: Optional[UploadFile],
    foto_registro_frente_file: Optional[UploadFile],
    foto_registro_reverso_file: Optional[UploadFile],
    gestor_cuenta_id: Optional[int],
    modified_by: str,
) -> schemas.Propietario:
    if repositories.get_propietario_by(db, data.tipo_persona_id, data.ruc):
        raise HTTPException(
            status_code=409,
            detail=f"El Propietario con documento {data.ruc} ya existe",
        )
    (
        foto_documento_frente_url,
        foto_documento_reverso_url,
        foto_perfil_url,
    ) = await check_files(
        foto_documento_frente_file,
        foto_documento_reverso_file,
        foto_perfil_file,
    )
    chofer = None
    if data.es_chofer:
        chofer = await create_or_edit_chofer_by_propietario(
            db,
            data,
            foto_documento_frente_chofer_file,
            foto_documento_reverso_chofer_file,
            foto_registro_frente_file,
            foto_registro_reverso_file,
            foto_perfil_url,
            gestor_cuenta_id,
            modified_by,
        )
    obj = repositories.create_propietario(
        db,
        data,
        gestor_cuenta_id,
        cast(str, foto_documento_frente_url),
        cast(str, foto_documento_reverso_url),
        cast(str, foto_perfil_url),
        chofer,
        modified_by,
    )
    update_propietario_contacto_list(
        db, data.contactos, obj, gestor_cuenta_id, modified_by
    )
    create_gestor_carga_propietario(db, obj, gestor_cuenta_id, data.alias, modified_by)
    return get_propietario_detail(obj, gestor_cuenta_id)


def get_propietario_by_id(db: Session, id: int) -> Propietario:
    obj = repositories.get_propietario_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Propietario no encontrado")
    return obj


def get_propietario_by_id_and_gestor_cuenta_id(
    db: Session, id: int, gestor_cuenta_id: Optional[int] = None
) -> schemas.Propietario:
    obj = repositories.get_propietario_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Propietario no encontrado")
    return get_propietario_detail(obj, gestor_cuenta_id)


async def edit_propietario(
    id: int,
    db: Session,
    data: schemas.PropietarioEditForm,
    foto_documento_frente_file: Optional[UploadFile],
    foto_documento_reverso_file: Optional[UploadFile],
    foto_perfil_file: Optional[UploadFile],
    foto_documento_frente_chofer_file: Optional[UploadFile],
    foto_documento_reverso_chofer_file: Optional[UploadFile],
    foto_registro_frente_file: Optional[UploadFile],
    foto_registro_reverso_file: Optional[UploadFile],
    gestor_cuenta_id: Optional[int],
    modified_by: str,
) -> schemas.Propietario:
    if data.tipo_persona_id and data.ruc:
        exists = repositories.get_propietario_by(db, data.tipo_persona_id, data.ruc)
        if exists and exists.id != id:
            raise HTTPException(
                status_code=409,
                detail=f"El Propietario con documento {data.ruc} ya existe",
            )
    (
        foto_documento_frente_url,
        foto_documento_reverso_url,
        foto_perfil_url,
    ) = await check_files(
        foto_documento_frente_file,
        foto_documento_reverso_file,
        foto_perfil_file,
    )
    to_edit_obj = get_propietario_by_id(db, id)
    chofer_id = to_edit_obj.chofer_id
    chofer = to_edit_obj.chofer
    if data.es_chofer:
        chofer_data = cast(schemas.PropietarioForm, data)
        chofer = await create_or_edit_chofer_by_propietario(
            db,
            chofer_data,
            foto_documento_frente_chofer_file,
            foto_documento_reverso_chofer_file,
            foto_registro_frente_file,
            foto_registro_reverso_file,
            foto_perfil_url,
            gestor_cuenta_id,
            modified_by,
            chofer_id,
        )
    else:
        disable_chofer_by_id(db, chofer_id, modified_by)
    obj = repositories.edit_propietario(
        to_edit_obj,
        db,
        data,
        foto_documento_frente_url,
        foto_documento_reverso_url,
        foto_perfil_url,
        chofer,
        modified_by,
    )
    update_propietario_contacto_list(
        db, data.contactos, obj, gestor_cuenta_id, modified_by
    )
    edit_gestor_carga_propietario(db, obj, gestor_cuenta_id, data.alias, modified_by)
    return get_propietario_detail(obj, gestor_cuenta_id)


def delete_propietario(
    db: Session, id: int, gestor_cuenta_id: Optional[int], modified_by: str
) -> schemas.Propietario:
    co = get_propietario_by_id(db, id)
    obj = repositories.delete_propietario(co, db, modified_by)
    return get_propietario_detail(obj, gestor_cuenta_id)


def change_propietario_status(
    db: Session, id: int, status: EstadoEnum, modified_by: str
) -> schemas.Camion:
    co = get_propietario_by_id(db, id)
    return repositories.change_propietario_status(co, db, status, modified_by)


def get_propietario_reports(db: Session) -> str:
    datalist = repositories.get_propietario_list(db)
    wb = Workbook()
    # get worksheet
    ws = wb.active

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Nombre"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Tipo de Persona"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "RUC"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Pais de Origen"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Gestor de Cuenta"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=6)
    title_cell.value = "Oficial de Cuenta"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Dirección"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=8)
    title_cell.value = "Ubicación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=9)
    title_cell.value = "Usuario creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=10)
    title_cell.value = "Fecha creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=11)
    title_cell.value = "Usuario modificación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=12)
    title_cell.value = "Fecha modificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.nombre

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.tipo_persona.descripcion

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.ruc

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.pais_origen.nombre

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = item.gestor_cuenta_nombre

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.oficial_cuenta_nombre

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = item.direccion if item.direccion else ""

        value_cell = ws.cell(row=row + 2, column=8)
        value_cell.value = f"{item.ciudad.nombre}/{item.ciudad.localidad.nombre}/{item.ciudad.localidad.pais.nombre_corto}"  # noqa

        value_cell = ws.cell(row=row + 2, column=9)
        value_cell.value = item.created_by

        value_cell = ws.cell(row=row + 2, column=10)
        value_cell.value = item.created_at

        value_cell = ws.cell(row=row + 2, column=11)
        value_cell.value = item.modified_by

        value_cell = ws.cell(row=row + 2, column=12)
        value_cell.value = item.modified_at

    ws.auto_filter.ref = ws.dimensions
    filename = "propietario_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
