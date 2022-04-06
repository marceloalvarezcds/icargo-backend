import os
from decimal import Decimal
from typing import List, Optional, Tuple

from fastapi import HTTPException  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories
from app.config import REPORTS_FOLDER
from app.enums import (
    InstrumentoViaEnum,
    LiquidacionEstadoEnum,
    LiquidacionEtapaEnum,
    OperacionEstadoEnum,
)
from app.models import Instrumento, InstrumentoVia, Liquidacion, TipoInstrumento
from app.schemas import InstrumentoForm, InstrumentoSaldoForm

from .banco import get_banco_by_id
from .caja import get_caja_by_id


def get_instrumento_by_id(db: Session, id: int) -> Instrumento:
    obj = repositories.get_instrumento_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Instrumento no encontrado")
    return obj


def get_instrumento_via_by_id(db: Session, id: int) -> InstrumentoVia:
    obj = repositories.get_instrumento_via_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Instrumento Vía no encontrado")
    return obj


def get_liquidacion_by_id(db: Session, id: int) -> Liquidacion:
    obj = repositories.get_liquidacion_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Liquidación no encontrada")
    return obj


def get_saldos_by_monto(
    db: Session,
    via: InstrumentoVia,
    caja_id: Optional[int],
    banco_id: Optional[int],
    liquidacion_id: int,
    monto: Decimal,
    modified_by: str,
) -> Tuple[Decimal, Decimal, Decimal, Decimal, Decimal]:
    liquidacion = get_liquidacion_by_id(db, liquidacion_id)
    credito = Decimal(0)
    debito = Decimal(0)
    provision = Decimal(0)
    saldo_confirmado = Decimal(0)
    saldo_provisional = Decimal(0)
    if caja_id and via.descripcion == InstrumentoViaEnum.CAJA.value:
        caja = get_caja_by_id(db, caja_id)
        if liquidacion.tipo_operacion_descripcion == "Cobro":
            credito = monto
        else:
            debito = monto
        saldo_confirmado = (
            credito - debito + (caja.saldo_confirmado if caja.saldo_confirmado else 0)
        )
        repositories.change_caja_saldos(caja, db, saldo_confirmado, modified_by)
    elif banco_id:
        banco = get_banco_by_id(db, banco_id)
        if liquidacion.tipo_operacion_descripcion == "Cobro":
            provision = monto
        else:
            provision = monto * -1
        saldo_confirmado = banco.saldo_confirmado if banco.saldo_confirmado else 0
        saldo_provisional = provision + (
            banco.saldo_provisional if banco.saldo_provisional else 0
        )
        repositories.change_banco_saldos(
            banco, db, saldo_confirmado, saldo_provisional, modified_by
        )
    return (
        credito,
        debito,
        saldo_confirmado,
        provision,
        saldo_provisional,
    )


def get_tipo_instrumento_by_descripcion(
    db: Session, descripcion: str
) -> TipoInstrumento:
    obj = repositories.get_tipo_instrumento_by_descripcion(db, descripcion)
    if not obj:
        raise HTTPException(status_code=404, detail="Tipo de Instrumento no encontrado")
    return obj


def finalizar_liquidacion(db: Session, liquidacion: Liquidacion, modified_by: str):
    is_saldo_cerrado = int(liquidacion.saldo_residual) == 0
    instrumentos: List[Instrumento] = liquidacion.instrumentos
    is_finalizado = is_saldo_cerrado
    if is_saldo_cerrado:
        for instrumento in instrumentos:
            if instrumento.operacion_estado == OperacionEstadoEnum.EMITIDO.value:
                is_finalizado = False
    if is_finalizado:
        liquidacion.etapa = LiquidacionEtapaEnum.FINALIZADO.value
        repositories.change_liquidacion_status(
            liquidacion, db, LiquidacionEstadoEnum.FINALIZADO, modified_by
        )


def create_instrumento(
    db: Session,
    liquidacion_id: int,
    data: InstrumentoForm,
    modified_by: str,
) -> Instrumento:
    via = get_instrumento_via_by_id(db, data.via_id)
    operacion_estado = OperacionEstadoEnum.EMITIDO
    if via.descripcion == InstrumentoViaEnum.CAJA.value:
        tipo_instrumento = get_tipo_instrumento_by_descripcion(db, "Efectivo")
        data.tipo_instrumento_id = tipo_instrumento.id
        operacion_estado = OperacionEstadoEnum.CONFIRMADO
    (
        credito,
        debito,
        saldo_confirmado,
        provision,
        saldo_provisional,
    ) = get_saldos_by_monto(
        db,
        via,
        data.caja_id,
        data.banco_id,
        liquidacion_id,
        data.monto,
        modified_by,
    )
    data.liquidacion_id = liquidacion_id
    instrumento = repositories.create_instrumento(
        db,
        InstrumentoSaldoForm(
            **data.dict(),
            operacion_estado=operacion_estado,
            credito=credito,
            debito=debito,
            saldo_confirmado=saldo_confirmado,
            provision=provision,
            saldo_provisional=saldo_provisional,
        ),
        modified_by,
    )
    finalizar_liquidacion(db, instrumento.liquidacion, modified_by)
    return instrumento


def edit_instrumento(
    id: int,
    db: Session,
    data: InstrumentoForm,
    modified_by: str,
) -> Instrumento:
    liquidacion_id = data.liquidacion_id if data.liquidacion_id else 0
    via = get_instrumento_via_by_id(db, data.via_id)
    to_edit_obj = get_instrumento_by_id(db, id)
    (
        credito,
        debito,
        saldo_confirmado,
        provision,
        saldo_provisional,
    ) = get_saldos_by_monto(
        db,
        via,
        data.caja_id,
        data.banco_id,
        liquidacion_id,
        data.monto,
        modified_by,
    )
    return repositories.edit_instrumento(
        to_edit_obj,
        db,
        InstrumentoSaldoForm(
            **data.dict(),
            credito=credito,
            debito=debito,
            saldo_confirmado=saldo_confirmado,
            provision=provision,
            saldo_provisional=saldo_provisional,
        ),
        modified_by,
    )


def delete_instrumento(db: Session, id: int, modified_by: str) -> Instrumento:
    obj = get_instrumento_by_id(db, id)
    return repositories.delete_instrumento(obj, db, modified_by)


def confirmar_instrumento(db: Session, id: int, modified_by: str) -> Instrumento:
    obj = get_instrumento_by_id(db, id)
    if obj.provision < 0:
        obj.debito = obj.provision * -1
    else:
        obj.credito = obj.provision
    banco = get_banco_by_id(db, obj.banco_id)
    saldo_confirmado = (
        obj.credito
        - obj.debito
        + (banco.saldo_confirmado if banco.saldo_confirmado else 0)
    )
    saldo_provisional = banco.saldo_provisional if banco.saldo_provisional else 0
    repositories.change_banco_saldos(
        banco, db, saldo_confirmado, saldo_provisional, modified_by
    )
    obj.provision = 0
    obj.saldo_confirmado = saldo_confirmado
    obj.saldo_provisional = saldo_provisional
    instrumento = repositories.change_instrumento_operacion_estado(
        obj, db, OperacionEstadoEnum.CONFIRMADO, modified_by
    )
    finalizar_liquidacion(db, instrumento.liquidacion, modified_by)
    return instrumento


def rechazar_instrumento(db: Session, id: int, modified_by: str) -> Instrumento:
    obj = get_instrumento_by_id(db, id)
    banco = get_banco_by_id(db, obj.banco_id)
    obj.provision_rechazada = obj.provision
    saldo_confirmado = banco.saldo_confirmado if banco.saldo_confirmado else 0
    saldo_provisional = (
        banco.saldo_provisional if banco.saldo_provisional else 0
    ) - obj.provision
    repositories.change_banco_saldos(
        banco, db, saldo_confirmado, saldo_provisional, modified_by
    )
    liquidacion = get_liquidacion_by_id(db, obj.liquidacion_id)
    repositories.change_liquidacion_status(
        liquidacion, db, LiquidacionEstadoEnum.SALDO_ABIERTO, modified_by
    )
    obj.provision = 0
    obj.saldo_confirmado = saldo_confirmado
    obj.saldo_provisional = saldo_provisional
    return repositories.change_instrumento_operacion_estado(
        obj, db, OperacionEstadoEnum.RECHAZADO, modified_by
    )


def get_instrumento_reports(db: Session) -> str:
    datalist = repositories.get_instrumento_list(db)
    wb = Workbook()
    ws = wb.active

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "ID"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Tipo de Contraparte"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "Nombre Contraparte"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Nº Contraparte"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Cobro/Pago"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=6)
    title_cell.value = "Moneda"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Estado"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=8)
    title_cell.value = "Crédito"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=9)
    title_cell.value = "Débito"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=10)
    title_cell.value = "Saldo Confirmado"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=11)
    title_cell.value = "Provisión"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=12)
    title_cell.value = "Saldo Total"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=13)
    title_cell.value = "Usuario creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=14)
    title_cell.value = "Fecha creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=15)
    title_cell.value = "Usuario modificación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=16)
    title_cell.value = "Fecha modificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.id

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.tipo_contraparte_descripcion

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.contraparte

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.contraparte_numero_documento

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = item.liquidacion.tipo_operacion_descripcion

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.moneda_nombre

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = item.operacion_estado

        value_cell = ws.cell(row=row + 2, column=8)
        value_cell.value = item.credito

        value_cell = ws.cell(row=row + 2, column=9)
        value_cell.value = item.debito

        value_cell = ws.cell(row=row + 2, column=10)
        value_cell.value = item.saldo_confirmado

        value_cell = ws.cell(row=row + 2, column=11)
        value_cell.value = item.provision

        value_cell = ws.cell(row=row + 2, column=12)
        value_cell.value = item.saldo_provisional

        value_cell = ws.cell(row=row + 2, column=13)
        value_cell.value = item.created_by

        value_cell = ws.cell(row=row + 2, column=14)
        value_cell.value = item.created_at

        value_cell = ws.cell(row=row + 2, column=15)
        value_cell.value = item.modified_by

        value_cell = ws.cell(row=row + 2, column=16)
        value_cell.value = item.modified_at

    ws.auto_filter.ref = ws.dimensions
    filename = "instrumento_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
