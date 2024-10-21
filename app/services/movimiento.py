import os
from typing import Union
from app.schemas.movimiento import EstadoCuentaMovimiento
from datetime import datetime
from http import HTTPStatus
from typing import List, Optional, cast

from fastapi import HTTPException  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories
from app.config import REPORTS_FOLDER
from app.enums import (
    MovimientoEstadoEnum,
    TipoContraparteEnum,
    TipoCuentaEnum,
    TipoMovimientoEnum,
)
from app.models import (
    Moneda,
    Movimiento,
    OrdenCarga,
    OrdenCargaAnticipoRetirado,
    OrdenCargaComplemento,
    OrdenCargaDescuento,
    TipoCuenta,
    TipoMovimiento,
    Propietario,
    Chofer,
    Proveedor,
    Remitente,
    Liquidacion,
)
from app.schemas import MovimientoFleteEditForm, MovimientoForm, MovimientoMermaEditForm, FacturaForm
from app.schemas.date_model import Date
from app.schemas.orden_carga import OrdenCargaEditForm
from app.schemas.rounded_decimal_model import RoundedDecimal
from app.services import seleccionable_service as service
from app.utils import get_flete_detalle, get_merma_detalle

from app.logger import logger

def get_orden_carga_by_movimiento(movimiento: Movimiento):
    oc = movimiento.orden_carga
    if not oc:
        raise HTTPException(
            status_code=HTTPStatus.NOT_FOUND,
            detail=f"No existe Orden de Carga para el movimiento {id}",
        )
    return oc


def get_moneda_by_id(db: Session, moneda_id: Optional[int]) -> Moneda:
    if not moneda_id:
        raise HTTPException(
            status_code=HTTPStatus.BAD_REQUEST,
            detail=f"No existe la moneda con ID {moneda_id}",
        )
    moneda = repositories.get_moneda_by_id(db, moneda_id)
    if not moneda:
        raise HTTPException(
            status_code=HTTPStatus.NOT_FOUND,
            detail=f"No existe la moneda con ID {moneda_id}",
        )
    return moneda


def get_movimiento_list(
    db: Session, gestor_carga_id: Optional[int]
) -> List[Movimiento]:
    if gestor_carga_id:
        return repositories.get_movimiento_list_by_gestor_carga_id(db, gestor_carga_id)
    return repositories.get_movimiento_list(db)


def get_movimiento_list_by_estado_cuenta(
    db: Session,
    tipo_contraparte_id: int,
    contraparte_id: int,
    contraparte: str,
    contraparte_numero_documento: str,
    estado: str,
    gestor_carga_id: Optional[int],
    punto_venta_id: Optional[int] = None,
) -> List[Movimiento]:
    if gestor_carga_id:
        return repositories.get_movimiento_list_by_contraparte_and_gestor_carga_id(
            db,
            tipo_contraparte_id,
            contraparte_id,
            contraparte,
            contraparte_numero_documento,
            estado,
            gestor_carga_id,
            punto_venta_id
        )
    return repositories.get_movimiento_list_by_contraparte(
        db,
        tipo_contraparte_id,
        contraparte_id,
        contraparte,
        contraparte_numero_documento,
        estado,
    )


def get_movimiento_list_by_liquidacion(
    db: Session,
    liquidacion_id: int,
    estado: str,
    gestor_carga_id: Optional[int],
) -> List[Movimiento]:
    if gestor_carga_id:
        return repositories.get_movimiento_list_by_liquidacion_and_gestor_carga_id(
            db,
            liquidacion_id,
            estado,
            gestor_carga_id,
        )
    return repositories.get_movimiento_list_by_liquidacion(db, liquidacion_id, estado)


def create_movimiento(
    db: Session,
    data: MovimientoForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Optional[Movimiento]:
    # Movimiento no puede tener monto 0
    if int(data.monto) != 0:
        gestor_id = gestor_carga_id if gestor_carga_id else data.gestor_carga_id
        if not gestor_id:
            raise HTTPException(
                status_code=HTTPStatus.CONFLICT, detail="Debe elegir un Gestor de carga"
            )
        return repositories.create_movimiento(db, data, gestor_id, modified_by)
    return None


def create_movimiento_by_anticipo(
    db: Session,
    anticipo: OrdenCargaAnticipoRetirado,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Optional[Movimiento]:
    propietario_contraparte = repositories.get_tipo_contraparte_by_descripcion(
        db, TipoContraparteEnum.PROPIETARIO.value
    )
    proveedor_contraparte = repositories.get_tipo_contraparte_by_descripcion(
        db, TipoContraparteEnum.PROVEEDOR.value
    )
    tipo_documento_relacionado = (
        repositories.get_tipo_documento_relacionado_by_descripcion(db, "OC")
    )
    tipo_cuenta = repositories.get_tipo_cuenta_by_descripcion(
        db, TipoCuentaEnum.VIAJES.value
    )
    tipo_movimiento = repositories.get_tipo_movimiento_by_descripcion(
        db, TipoMovimientoEnum.ANTICIPO.value
    )
    if (
        not propietario_contraparte
        or not proveedor_contraparte
        or not tipo_documento_relacionado
        or not tipo_cuenta
        or not tipo_movimiento
    ):
        raise HTTPException(
            status_code=HTTPStatus.NOT_FOUND,
            detail="Tipo de contraparte, doc relacionado, cuenta o movimiento no existe",
        )
    create_movimiento(
        db,
        MovimientoForm(
            orden_carga_id=anticipo.orden_carga_id,
            tipo_contraparte_id=proveedor_contraparte.id,
            contraparte=anticipo.proveedor_nombre,
            contraparte_numero_documento=anticipo.punto_venta.proveedor.numero_documento,
            tipo_documento_relacionado_id=tipo_documento_relacionado.id,
            numero_documento_relacionado=anticipo.orden_carga_id,
            cuenta_id=tipo_cuenta.id,
            tipo_movimiento_id=tipo_movimiento.id,
            estado=MovimientoEstadoEnum.PENDIENTE,
            detalle=anticipo.detalle,
            monto=anticipo.monto_retirado,
            moneda_id=anticipo.moneda_id,
            tipo_cambio_moneda=1,  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en Descuento  # noqa
            fecha_cambio_moneda=datetime.now(),
            anticipo_id=anticipo.id,
            proveedor_id=anticipo.punto_venta.proveedor_id,
            tipo_movimiento_info=anticipo.concepto
        ),
        gestor_carga_id,
        modified_by,
    )
    return create_movimiento(
        db,
        MovimientoForm(
            orden_carga_id=anticipo.orden_carga_id,
            tipo_contraparte_id=propietario_contraparte.id,
            contraparte=anticipo.orden_carga.camion_propietario_nombre,
            contraparte_numero_documento=anticipo.orden_carga.camion.propietario_ruc,
            tipo_documento_relacionado_id=tipo_documento_relacionado.id,
            numero_documento_relacionado=anticipo.orden_carga_id,
            cuenta_id=tipo_cuenta.id,
            tipo_movimiento_id=tipo_movimiento.id,
            estado=MovimientoEstadoEnum.PENDIENTE,
            detalle=anticipo.detalle,
            monto=-anticipo.monto_retirado,
            moneda_id=anticipo.moneda_id,
            tipo_cambio_moneda=1,  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en anticipos  # noqa
            fecha_cambio_moneda=datetime.now(),
            anticipo_id=anticipo.id,
            propietario_id=anticipo.orden_carga.camion.propietario_id,
            tipo_movimiento_info=anticipo.concepto
        ),
        gestor_carga_id,
        modified_by,
    )


def create_movimiento_by_flete(
    db: Session,
    orden_carga: OrdenCarga,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Optional[Movimiento]:
    propietario_contraparte = repositories.get_tipo_contraparte_by_descripcion(
        db, TipoContraparteEnum.PROPIETARIO.value
    )
    remitente_contraparte = repositories.get_tipo_contraparte_by_descripcion(
        db, TipoContraparteEnum.REMITENTE.value
    )
    tipo_documento_relacionado = (
        repositories.get_tipo_documento_relacionado_by_descripcion(db, "OC")
    )
    tipo_cuenta = repositories.get_tipo_cuenta_by_descripcion(
        db, TipoCuentaEnum.VIAJES.value
    )
    tipo_movimiento = repositories.get_tipo_movimiento_by_descripcion(
        db, TipoMovimientoEnum.FLETE.value
    )
    if (
        not propietario_contraparte
        or not remitente_contraparte
        or not tipo_documento_relacionado
        or not tipo_cuenta
        or not tipo_movimiento
    ):
        raise HTTPException(
            status_code=HTTPStatus.NOT_FOUND,
            detail="Tipo de contraparte, doc relacionado, cuenta o movimiento no existe",
        )
    create_movimiento(
        db,
        MovimientoForm(
            orden_carga_id=orden_carga.id,
            tipo_contraparte_id=remitente_contraparte.id,
            contraparte=orden_carga.flete.remitente_nombre,
            contraparte_numero_documento=orden_carga.flete.remitente.numero_documento,
            tipo_documento_relacionado_id=tipo_documento_relacionado.id,
            numero_documento_relacionado=orden_carga.id,
            cuenta_id=tipo_cuenta.id,
            tipo_movimiento_id=tipo_movimiento.id,
            estado=MovimientoEstadoEnum.PENDIENTE,
            detalle=orden_carga.flete_gestor_carga_detalle,
            monto=-orden_carga.resultado_gestor_carga_total_flete,
            moneda_id=orden_carga.flete.condicion_gestor_cuenta_moneda_id,
            tipo_cambio_moneda=1,  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en FLETE  # noqa
            fecha_cambio_moneda=datetime.now(),
            remitente_id=orden_carga.flete.remitente_id,
            tipo_movimiento_info=tipo_movimiento.descripcion,
        ),
        gestor_carga_id,
        modified_by,
    )
    return create_movimiento(
        db,
        MovimientoForm(
            orden_carga_id=orden_carga.id,
            tipo_contraparte_id=propietario_contraparte.id,
            contraparte=orden_carga.camion_propietario_nombre,
            contraparte_numero_documento=orden_carga.camion.propietario_ruc,
            tipo_documento_relacionado_id=tipo_documento_relacionado.id,
            numero_documento_relacionado=orden_carga.id,
            cuenta_id=tipo_cuenta.id,
            tipo_movimiento_id=tipo_movimiento.id,
            estado=MovimientoEstadoEnum.PENDIENTE,
            detalle=orden_carga.flete_propietario_detalle,
            monto=orden_carga.resultado_propietario_total_flete,
            moneda_id=orden_carga.flete.condicion_propietario_moneda_id,
            tipo_cambio_moneda=1,  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en FLETE  # noqa
            fecha_cambio_moneda=datetime.now(),
            propietario_id=orden_carga.camion.propietario_id,
            tipo_movimiento_info=tipo_movimiento.descripcion
        ),
        gestor_carga_id,
        modified_by,
    )


def create_movimiento_by_complemento(
    db: Session,
    complemento: OrdenCargaComplemento,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Optional[Movimiento]:
    propietario_contraparte = repositories.get_tipo_contraparte_by_descripcion(
        db, TipoContraparteEnum.PROPIETARIO.value
    )
    remitente_contraparte = repositories.get_tipo_contraparte_by_descripcion(
        db, TipoContraparteEnum.REMITENTE.value
    )
    tipo_documento_relacionado = (
        repositories.get_tipo_documento_relacionado_by_descripcion(db, "OC")
    )
    tipo_cuenta = repositories.get_tipo_cuenta_by_descripcion(
        db, TipoCuentaEnum.VIAJES.value
    )
    tipo_movimiento = repositories.get_tipo_movimiento_by_descripcion(
        db, TipoMovimientoEnum.COMPLEMENTO.value
    )
    if (
        not propietario_contraparte
        or not remitente_contraparte
        or not tipo_documento_relacionado
        or not tipo_cuenta
        or not tipo_movimiento
    ):
        raise HTTPException(
            status_code=HTTPStatus.NOT_FOUND,
            detail="Tipo de contraparte, doc relacionado, cuenta o movimiento no existe",
        )
    if complemento.habilitar_cobro_remitente:
        create_movimiento(
            db,
            MovimientoForm(
                orden_carga_id=complemento.orden_carga_id,
                tipo_contraparte_id=remitente_contraparte.id,
                contraparte=complemento.orden_carga.flete.remitente_nombre,
                contraparte_numero_documento=complemento.orden_carga.flete.remitente.numero_documento,  # noqa
                tipo_documento_relacionado_id=tipo_documento_relacionado.id,
                numero_documento_relacionado=complemento.orden_carga_id,
                cuenta_id=tipo_cuenta.id,
                tipo_movimiento_id=tipo_movimiento.id,
                estado=MovimientoEstadoEnum.PENDIENTE,
                detalle=complemento.remitente_detalle,
                monto=-complemento.remitente_monto,
                moneda_id=complemento.remitente_moneda_id,
                tipo_cambio_moneda=1,  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en Complemento  # noqa
                fecha_cambio_moneda=datetime.now(),
                complemento_id=complemento.id,
                remitente_id=complemento.orden_carga.flete.remitente_id,
                tipo_movimiento_info=complemento.concepto_descripcion,
            ),
            gestor_carga_id,
            modified_by,
        )
    return create_movimiento(
        db,
        MovimientoForm(
            orden_carga_id=complemento.orden_carga_id,
            tipo_contraparte_id=propietario_contraparte.id,
            contraparte=complemento.orden_carga.camion_propietario_nombre,
            contraparte_numero_documento=complemento.orden_carga.camion.propietario_ruc,
            tipo_documento_relacionado_id=tipo_documento_relacionado.id,
            numero_documento_relacionado=complemento.orden_carga_id,
            cuenta_id=tipo_cuenta.id,
            tipo_movimiento_id=tipo_movimiento.id,
            estado=MovimientoEstadoEnum.PENDIENTE,
            detalle=complemento.propietario_detalle,
            monto=complemento.propietario_monto,
            moneda_id=complemento.propietario_moneda_id,
            tipo_cambio_moneda=1,  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en Complemento  # noqa
            fecha_cambio_moneda=datetime.now(),
            complemento_id=complemento.id,
            propietario_id=complemento.orden_carga.camion.propietario_id,
            tipo_movimiento_info=complemento.concepto_descripcion,
        ),
        gestor_carga_id,
        modified_by,
    )


def create_movimiento_by_descuento(
    db: Session,
    descuento: OrdenCargaDescuento,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Optional[Movimiento]:
    propietario_contraparte = repositories.get_tipo_contraparte_by_descripcion(
        db, TipoContraparteEnum.PROPIETARIO.value
    )
    proveedor_contraparte = repositories.get_tipo_contraparte_by_descripcion(
        db, TipoContraparteEnum.PROVEEDOR.value
    )
    tipo_documento_relacionado = (
        repositories.get_tipo_documento_relacionado_by_descripcion(db, "OC")
    )
    tipo_cuenta = repositories.get_tipo_cuenta_by_descripcion(
        db, TipoCuentaEnum.VIAJES.value
    )
    tipo_movimiento = repositories.get_tipo_movimiento_by_descripcion(
        db, TipoMovimientoEnum.DESCUENTO.value
    )
    if (
        not propietario_contraparte
        or not proveedor_contraparte
        or not tipo_documento_relacionado
        or not tipo_cuenta
        or not tipo_movimiento
    ):
        raise HTTPException(
            status_code=HTTPStatus.NOT_FOUND,
            detail="Tipo de contraparte, doc relacionado, cuenta o movimiento no existe ooooo",
        )
    if descuento.habilitar_pago_proveedor and descuento.proveedor:
        create_movimiento(
            db,
            MovimientoForm(
                orden_carga_id=descuento.orden_carga_id,
                tipo_contraparte_id=proveedor_contraparte.id,
                contraparte=descuento.proveedor_nombre,
                contraparte_numero_documento=descuento.proveedor.numero_documento,
                tipo_documento_relacionado_id=tipo_documento_relacionado.id,
                numero_documento_relacionado=descuento.orden_carga_id,
                cuenta_id=tipo_cuenta.id,
                tipo_movimiento_id=tipo_movimiento.id,
                estado=MovimientoEstadoEnum.PENDIENTE,
                detalle=descuento.proveedor_detalle,
                monto=descuento.proveedor_monto,
                moneda_id=descuento.proveedor_moneda_id,
                tipo_cambio_moneda=1,  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en Descuento  # noqa
                fecha_cambio_moneda=datetime.now(),
                descuento_id=descuento.id,
                proveedor_id=descuento.proveedor_id,
                tipo_movimiento_info=descuento.concepto_descripcion,
            ),
            gestor_carga_id,
            modified_by,
        )
    return create_movimiento(
        db,
        MovimientoForm(
            orden_carga_id=descuento.orden_carga_id,
            tipo_contraparte_id=propietario_contraparte.id,
            contraparte=descuento.orden_carga.camion_propietario_nombre,
            contraparte_numero_documento=descuento.orden_carga.camion.propietario_ruc,
            tipo_documento_relacionado_id=tipo_documento_relacionado.id,
            numero_documento_relacionado=descuento.orden_carga_id,
            cuenta_id=tipo_cuenta.id,
            tipo_movimiento_id=tipo_movimiento.id,
            estado=MovimientoEstadoEnum.PENDIENTE,
            detalle=descuento.propietario_detalle,
            monto=-descuento.propietario_monto,
            moneda_id=descuento.propietario_moneda_id,
            tipo_cambio_moneda=1,  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en Descuento  # noqa
            fecha_cambio_moneda=datetime.now(),
            descuento_id=descuento.id,
            propietario_id=descuento.orden_carga.camion.propietario_id,
            tipo_movimiento_info=descuento.concepto_descripcion,
        ),
        gestor_carga_id,
        modified_by,
    )


def create_movimiento_by_merma(
    db: Session,
    orden_carga: OrdenCarga,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Optional[Movimiento]:
    propietario_contraparte = repositories.get_tipo_contraparte_by_descripcion(
        db, TipoContraparteEnum.PROPIETARIO.value
    )
    remitente_contraparte = repositories.get_tipo_contraparte_by_descripcion(
        db, TipoContraparteEnum.REMITENTE.value
    )
    tipo_documento_relacionado = (
        repositories.get_tipo_documento_relacionado_by_descripcion(db, "OC")
    )
    tipo_cuenta = repositories.get_tipo_cuenta_by_descripcion(
        db, TipoCuentaEnum.VIAJES.value
    )
    tipo_movimiento = repositories.get_tipo_movimiento_by_descripcion(
        db, TipoMovimientoEnum.MERMA.value
    )
    if (
        not propietario_contraparte
        or not remitente_contraparte
        or not tipo_documento_relacionado
        or not tipo_cuenta
        or not tipo_movimiento
    ):
        raise HTTPException(
            status_code=HTTPStatus.NOT_FOUND,
            detail="Tipo de contraparte, doc relacionado, cuenta o movimiento no existe",
        )
    create_movimiento(
        db,
        MovimientoForm(
            orden_carga_id=orden_carga.id,
            tipo_contraparte_id=remitente_contraparte.id,
            contraparte=orden_carga.flete.remitente_nombre,
            contraparte_numero_documento=orden_carga.flete.remitente.numero_documento,
            tipo_documento_relacionado_id=tipo_documento_relacionado.id,
            numero_documento_relacionado=orden_carga.id,
            cuenta_id=tipo_cuenta.id,
            tipo_movimiento_id=tipo_movimiento.id,
            estado=MovimientoEstadoEnum.PENDIENTE,
            detalle=orden_carga.merma_gestor_carga_detalle,
            monto=orden_carga.resultado_gestor_carga_merma_valor_total,
            moneda_id=orden_carga.flete.condicion_gestor_cuenta_moneda_id,
            tipo_cambio_moneda=1,  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en FLETE  # noqa
            fecha_cambio_moneda=datetime.now(),
            remitente_id=orden_carga.flete.remitente_id,
            tipo_movimiento_info=tipo_movimiento.descripcion,
        ),
        gestor_carga_id,
        modified_by,
    )
    return create_movimiento(
        db,
        MovimientoForm(
            orden_carga_id=orden_carga.id,
            tipo_contraparte_id=propietario_contraparte.id,
            contraparte=orden_carga.camion_propietario_nombre,
            contraparte_numero_documento=orden_carga.camion.propietario_ruc,
            tipo_documento_relacionado_id=tipo_documento_relacionado.id,
            numero_documento_relacionado=orden_carga.id,
            cuenta_id=tipo_cuenta.id,
            tipo_movimiento_id=tipo_movimiento.id,
            estado=MovimientoEstadoEnum.PENDIENTE,
            detalle=orden_carga.merma_propietario_detalle,
            monto=-orden_carga.resultado_propietario_merma_valor_total,
            moneda_id=orden_carga.flete.condicion_propietario_moneda_id,
            tipo_cambio_moneda=1,  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en FLETE  # noqa
            fecha_cambio_moneda=datetime.now(),
            propietario_id=orden_carga.camion.propietario_id,
            tipo_movimiento_info=tipo_movimiento.descripcion,
        ),
        gestor_carga_id,
        modified_by,
    )


def create_movimiento_by_tipo_documento_relacionado_otro(
    db: Session,
    data: MovimientoForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Optional[Movimiento]:
    tipo_contraparte = repositories.get_tipo_contraparte_by_descripcion(
        db, TipoContraparteEnum.OTRO.value
    )
    tipo_documento_relacionado = (
        repositories.get_tipo_documento_relacionado_by_descripcion(db, "Otro")
    )
    tipo_cuenta = (
        service.get_by_id(TipoCuenta, db, data.cuenta_id)
        if data.cuenta_id
        else repositories.get_tipo_cuenta_by_descripcion(
            db, TipoCuentaEnum.VIAJES.value
        )
    )
    tipo_movimiento = (
        service.get_by_id(TipoMovimiento, db, data.tipo_movimiento_id)
        if data.tipo_movimiento_id
        else repositories.get_tipo_movimiento_by_descripcion(
            db, TipoMovimientoEnum.OTRO.value
        )
    )
    if (
        not tipo_contraparte
        or not tipo_documento_relacionado
        or not tipo_cuenta
        or not tipo_movimiento
    ):
        raise HTTPException(
            status_code=HTTPStatus.NOT_FOUND,
            detail="Tipo de contraparte, doc relacionado, cuenta o movimiento no existe",
        )
    exists = None
    if data.es_creacion_contraparte:
        exists = repositories.get_contraparte_by_contraparte_and_tipo_contraparte_id(
            db, data.contraparte, tipo_contraparte.id
        )
    if exists:
        raise HTTPException(
            status_code=HTTPStatus.CONFLICT,
            detail=f"La contraparte {data.contraparte} ya existe",
        )
    numero_documento_relacionado = (
        repositories.get_movimiento_count_by_tipo_documento_relacionado_id(
            db, tipo_documento_relacionado.id
        )
        + 1
    )
    fecha = cast(Date, datetime.now().strftime("%Y-%m-%dT%H:%M:%S.%fZ"))
    data.tipo_documento_relacionado_id = tipo_documento_relacionado.id
    data.numero_documento_relacionado = numero_documento_relacionado
    data.cuenta_id = tipo_cuenta.id
    data.tipo_movimiento_id = tipo_movimiento.id
    data.es_editable = True
    data.tipo_cambio_moneda = RoundedDecimal(
        1
    )  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en Descuento  # noqa
    data.fecha_cambio_moneda = fecha
    if not data.fecha:
        data.fecha = fecha
    if not data.es_cobro:
        data.monto = data.monto * -1  # type: ignore
    return create_movimiento(
        db,
        data,
        gestor_carga_id,
        modified_by,
    )


def create_movimiento_by_conciliacion_oc(
    db: Session,
    orden_carga: OrdenCarga,
    gestor_carga_id: Optional[int],
    modified_by: str,
):
    for complemento in orden_carga.complementos:
        create_movimiento_by_complemento(db, complemento, gestor_carga_id, modified_by)
    for descuento in orden_carga.descuentos:
        create_movimiento_by_descuento(db, descuento, gestor_carga_id, modified_by)
    create_movimiento_by_flete(db, orden_carga, gestor_carga_id, modified_by)
    create_movimiento_by_merma(db, orden_carga, gestor_carga_id, modified_by)


def get_movimiento_by_id(db: Session, id: int) -> Movimiento:
    obj = repositories.get_movimiento_by_id(db, id)
    if not obj:
        raise HTTPException(
            status_code=HTTPStatus.NOT_FOUND, detail="Movimiento no encontrado"
        )
    return obj


def edit_movimiento(
    id: int,
    db: Session,
    data: MovimientoForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Movimiento:
    gestor_id = gestor_carga_id if gestor_carga_id else data.gestor_carga_id
    if not gestor_id:
        raise HTTPException(
            status_code=HTTPStatus.CONFLICT, detail="Debe elegir un Gestor de carga"
        )
    to_edit_obj = get_movimiento_by_id(db, id)
    if not data.es_cobro:
        data.monto = data.monto * -1  # type: ignore
    return repositories.edit_movimiento(to_edit_obj, db, data, gestor_id, modified_by)


def edit_movimiento_by_gestor_flete(
    id: int,
    db: Session,
    data: MovimientoFleteEditForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Optional[Movimiento]:
    if not gestor_carga_id:
        raise HTTPException(
            status_code=HTTPStatus.CONFLICT, detail="Debe elegir un Gestor de carga"
        )
    to_edit_obj = get_movimiento_by_id(db, id)
    oc = get_orden_carga_by_movimiento(to_edit_obj)
    orden = repositories.edit_orden_carga_by_movimiento(
        oc,
        db,
        OrdenCargaEditForm(
            condicion_gestor_carga_moneda_id=data.moneda_id,
            condicion_gestor_carga_tarifa=data.tarifa,
        ),
        gestor_carga_id,
        modified_by,
    )
    moneda_id = data.moneda_id
    moneda = get_moneda_by_id(db, moneda_id)
    monto = orden.resultado_gestor_carga_total_flete * -1
    unidad = orden.flete.condicion_gestor_cuenta_unidad
    tarifa = data.tarifa if data.tarifa else orden.flete_tarifa_gestor_carga
    detalle = get_flete_detalle(orden, tarifa, moneda, unidad)
    return repositories.edit_monto_movimiento(
        to_edit_obj, db, monto, detalle, moneda_id, gestor_carga_id, modified_by
    )


def edit_movimiento_by_gestor_merma(
    id: int,
    db: Session,
    data: MovimientoMermaEditForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Optional[Movimiento]:
    if not gestor_carga_id:
        raise HTTPException(
            status_code=HTTPStatus.CONFLICT, detail="Debe elegir un Gestor de carga"
        )
    to_edit_obj = get_movimiento_by_id(db, id)
    oc = get_orden_carga_by_movimiento(to_edit_obj)
    orden = repositories.edit_orden_carga_by_movimiento(
        oc,
        db,
        OrdenCargaEditForm(
            merma_gestor_carga_es_porcentual=data.es_porcentual,
            merma_gestor_carga_moneda_id=data.moneda_id,
            merma_gestor_carga_tolerancia=data.tolerancia,
            merma_gestor_carga_valor=data.valor,
        ),
        gestor_carga_id,
        modified_by,
    )
    moneda_id = data.moneda_id
    moneda = get_moneda_by_id(db, moneda_id)
    monto = orden.resultado_gestor_carga_merma_valor_total
    valor = data.valor if data.valor else orden.merma_gestor_carga_valor
    tolerancia = data.tolerancia if data.valor else orden.merma_gestor_carga_tolerancia
    es_porcentual = (
        data.es_porcentual if data.valor else orden.merma_gestor_carga_es_porcentual
    )
    detalle = get_merma_detalle(
        orden,
        valor,
        tolerancia,
        es_porcentual,
        moneda,
        orden.flete.merma_gestor_cuenta_unidad,
    )
    return repositories.edit_monto_movimiento(
        to_edit_obj, db, monto, detalle, moneda_id, gestor_carga_id, modified_by
    )


def edit_movimiento_by_propietario_flete(
    id: int,
    db: Session,
    data: MovimientoFleteEditForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Optional[Movimiento]:
    if not gestor_carga_id:
        raise HTTPException(
            status_code=HTTPStatus.CONFLICT, detail="Debe elegir un Gestor de carga"
        )
    to_edit_obj = get_movimiento_by_id(db, id)
    oc = get_orden_carga_by_movimiento(to_edit_obj)
    orden = repositories.edit_orden_carga_by_movimiento(
        oc,
        db,
        OrdenCargaEditForm(
            condicion_propietario_moneda_id=data.moneda_id,
            condicion_propietario_tarifa=data.tarifa,
        ),
        gestor_carga_id,
        modified_by,
    )
    moneda_id = data.moneda_id
    moneda = get_moneda_by_id(db, moneda_id)
    monto = orden.resultado_propietario_total_flete
    unidad = orden.flete.condicion_propietario_unidad
    tarifa = data.tarifa if data.tarifa else orden.flete_tarifa
    detalle = get_flete_detalle(orden, tarifa, moneda, unidad)
    return repositories.edit_monto_movimiento(
        to_edit_obj, db, monto, detalle, moneda_id, gestor_carga_id, modified_by
    )


def edit_movimiento_by_propietario_merma(
    id: int,
    db: Session,
    data: MovimientoMermaEditForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Optional[Movimiento]:
    if not gestor_carga_id:
        raise HTTPException(
            status_code=HTTPStatus.CONFLICT, detail="Debe elegir un Gestor de carga"
        )
    to_edit_obj = get_movimiento_by_id(db, id)
    oc = get_orden_carga_by_movimiento(to_edit_obj)
    orden = repositories.edit_orden_carga_by_movimiento(
        oc,
        db,
        OrdenCargaEditForm(
            merma_propietario_es_porcentual=data.es_porcentual,
            merma_propietario_moneda_id=data.moneda_id,
            merma_propietario_tolerancia=data.tolerancia,
            merma_propietario_valor=data.valor,
        ),
        gestor_carga_id,
        modified_by,
    )
    moneda_id = data.moneda_id
    moneda = get_moneda_by_id(db, moneda_id)
    monto = orden.resultado_propietario_merma_valor_total * -1
    valor = data.valor if data.valor else orden.merma_propietario_valor
    tolerancia = data.tolerancia if data.valor else orden.merma_propietario_tolerancia
    es_porcentual = (
        data.es_porcentual if data.valor else orden.merma_propietario_es_porcentual
    )
    detalle = get_merma_detalle(
        orden,
        valor,
        tolerancia,
        es_porcentual,
        moneda,
        orden.flete.merma_propietario_unidad,
    )
    return repositories.edit_monto_movimiento(
        to_edit_obj, db, monto, detalle, moneda_id, gestor_carga_id, modified_by
    )


def delete_movimiento(db: Session, id: int, modified_by: str) -> Movimiento:
    co = get_movimiento_by_id(db, id)
    co.liquidacion_id = None
    return repositories.delete_movimiento(co, db, modified_by)


def get_movimiento_reports_by_contraparte(
    db: Session,
    tipo_contraparte_id: int,
    contraparte_id: int,
    contraparte: str,
    contraparte_numero_documento: str,
    estado: str,
    gestor_carga_id: Optional[int] = None,
) -> str:
    datalist = []
    if gestor_carga_id:
        datalist = repositories.get_movimiento_list_for_reports_by_contraparte_and_gestor_carga_id(  # noqa: B950
            db,
            tipo_contraparte_id,
            contraparte_id,
            contraparte,
            contraparte_numero_documento,
            estado,
            gestor_carga_id,
        )
    datalist = repositories.get_movimiento_list_for_reports_by_contraparte(
        db,
        tipo_contraparte_id,
        contraparte_id,
        contraparte,
        contraparte_numero_documento,
        estado,
    )
    return generate_movimiento_reports(datalist)


def get_movimiento_reports_by_gestor_carga_id(
    db: Session, gestor_carga_id: Optional[int] = None
):
    datalist = get_movimiento_list(db, gestor_carga_id)
    return generate_movimiento_reports(datalist, True)


def get_movimiento_reports(
    db: Session,
    liquidacion_id: Optional[int] = None,
    estado: Optional[str] = None,
) -> str:
    datalist = []
    if estado and liquidacion_id:
        datalist = repositories.get_movimiento_list_for_reports_by_liquidacion_id(
            db, liquidacion_id, estado
        )
    else:
        datalist = repositories.get_movimiento_list(db)
    return generate_movimiento_reports(datalist)


def generate_movimiento_reports(
    datalist: List[Union[EstadoCuentaMovimiento, Movimiento]], is_for_listado: bool = False, is_with_saldos: bool = False
) -> str:
    wb = Workbook()
    ws = wb.active
    i = 0

    if is_for_listado:
        title_cell = ws.cell(row=1, column=(i := i + 1))
        title_cell.value = "Nº de Movimiento"
        title_cell.font = Font(bold=True)

        title_cell = ws.cell(row=1, column=(i := i + 1))
        title_cell.value = "Nº de Liquidación"
        title_cell.font = Font(bold=True)

        title_cell = ws.cell(row=1, column=(i := i + 1))
        title_cell.value = "Fecha de creación Liq."
        title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Estado"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Tipo Contraparte"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Contraparte"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Nº de Doc " + "Contraparte" if is_for_listado else "Fiscal"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Fecha"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Cuenta"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Concepto"
    title_cell.font = Font(bold=True)

    if is_for_listado:
        title_cell = ws.cell(row=1, column=(i := i + 1))
        title_cell.value = "Proveedor"
        title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Punto de Venta"
    title_cell.font = Font(bold=True)

    if not is_for_listado:
        title_cell = ws.cell(row=1, column=(i := i + 1))
        title_cell.value = "Nº de Mov."
        title_cell.font = Font(bold=True)

        title_cell = ws.cell(row=1, column=(i := i + 1))
        title_cell.value = "Nº de Liq."
        title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Nº Doc Relac."
    title_cell.font = Font(bold=True)

    if not is_with_saldos:
        title_cell = ws.cell(row=1, column=(i := i + 1))
        title_cell.value = "Monto"
        title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Detalle"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=(i := i + 1))
    title_cell.value = "Usuario"
    title_cell.font = Font(bold=True)

    if is_with_saldos:

        title_cell = ws.cell(row=1, column=(i := i + 1))
        title_cell.value = "Pendiente"
        title_cell.font = Font(bold=True)

        title_cell = ws.cell(row=1, column=(i := i + 1))
        title_cell.value = "Confirmado"
        title_cell.font = Font(bold=True)

        title_cell = ws.cell(row=1, column=(i := i + 1))
        title_cell.value = "Pagos/Cobros"
        title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        i = 0

        if is_for_listado:
            value_cell = ws.cell(row=row + 2, column=(i := i + 1))
            value_cell.value = item.id

            value_cell = ws.cell(row=row + 2, column=(i := i + 1))
            value_cell.value = item.liquidacion_id

            value_cell = ws.cell(row=row + 2, column=(i := i + 1))
            value_cell.value = item.liquidacion_fecha_creacion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.estado

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.tipo_contraparte_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.contraparte

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.contraparte_numero_documento

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.fecha if item.fecha else item.created_at

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.cuenta_codigo_descripcion

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.concepto

        if is_for_listado:
            value_cell = ws.cell(row=row + 2, column=(i := i + 1))
            value_cell.value = item.proveedor_nombre

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.punto_venta_nombre

        if not is_for_listado:
            value_cell = ws.cell(row=row + 2, column=(i := i + 1))
            value_cell.value = item.id

            value_cell = ws.cell(row=row + 2, column=(i := i + 1))
            value_cell.value = item.liquidacion_id

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.numero_documento_relacionado

        if not is_with_saldos:
            value_cell = ws.cell(row=row + 2, column=(i := i + 1))
            value_cell.value = item.monto

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.detalle

        value_cell = ws.cell(row=row + 2, column=(i := i + 1))
        value_cell.value = item.created_by

        if is_with_saldos:
            value_cell = ws.cell(row=row + 2, column=(i := i + 1))
            value_cell.value = item.pendiente

            value_cell = ws.cell(row=row + 2, column=(i := i + 1))
            value_cell.value = item.confirmado

            value_cell = ws.cell(row=row + 2, column=(i := i + 1))
            value_cell.value = item.finalizado


    ws.auto_filter.ref = ws.dimensions
    filename = "movimiento_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename


def get_all_movimiento_list_by_estado_cuenta(
    db: Session,
    tipo_contraparte_id: int,
    contraparte_id: int,
    contraparte: str,
    contraparte_numero_documento: str,
    gestor_carga_id: Optional[int],
    punto_venta_id: Optional[int] = None,
) -> List[EstadoCuentaMovimiento]:
    if gestor_carga_id:
        return repositories.get_all_movimiento_list_by_contraparte_and_gestor_carga_id(
            db,
            tipo_contraparte_id,
            contraparte_id,
            contraparte,
            contraparte_numero_documento,
            gestor_carga_id,
            punto_venta_id,
        )

    return repositories.get_all_movimiento_estado_cuenta_list_by_contraparte(
        db,
        tipo_contraparte_id,
        contraparte_id,
        contraparte,
        contraparte_numero_documento
    )


def get_movimiento_estado_cuenta_reports_by_contraparte(
    db: Session,
    tipo_contraparte_id: int,
    contraparte_id: int,
    contraparte: str,
    contraparte_numero_documento: str,
    gestor_carga_id: Optional[int] = None,
    punto_venta_id: Optional[int] = None,
) -> str:
    datalist = []
    if gestor_carga_id:
        datalist = repositories.get_all_movimiento_list_by_contraparte_and_gestor_carga_id(  # noqa: B950
            db,
            tipo_contraparte_id,
            contraparte_id,
            contraparte,
            contraparte_numero_documento,
            gestor_carga_id,
            punto_venta_id
        )
    datalist = repositories.get_all_movimiento_estado_cuenta_list_by_contraparte(
        db,
        tipo_contraparte_id,
        contraparte_id,
        contraparte,
        contraparte_numero_documento,
        punto_venta_id
    )
    return generate_movimiento_reports(datalist, True, True)



def create_movimiento_by_factura(
    db: Session,
    factura: FacturaForm,
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> Optional[Movimiento]:

    tipo_contraparte = repositories.get_tipo_comprobante_by_id(
        db, factura.tipo_contraparte_id
    )
    tipo_documento_relacionado = (
        repositories.get_tipo_documento_relacionado_by_descripcion(db, "OC")
    )
    tipo_cuenta = repositories.get_tipo_cuenta_by_descripcion(
        db, TipoCuentaEnum.VIAJES.value
    )
    tipo_movimiento = repositories.get_tipo_movimiento_by_descripcion(
        db, TipoMovimientoEnum.FISCAL.value
    )
    if (
        not tipo_contraparte
        or not tipo_documento_relacionado
        or not tipo_cuenta
        or not tipo_movimiento
    ):
        raise HTTPException(
            status_code=HTTPStatus.NOT_FOUND,
            detail="Tipo de contraparte, doc relacionado, cuenta o movimiento no existe",
        )

    chofer_id = None
    proveedor_id = None
    propietario_id = None
    remitente_id = None
    punto_venta_id = None

    if tipo_contraparte.descripcion == TipoContraparteEnum.CHOFER.value:
        contraparte = service.get_by_id(Chofer, db, factura.contraparte_id)
        chofer_id = contraparte.id

    if tipo_contraparte.descripcion == TipoContraparteEnum.PROVEEDOR.value:
        contraparte = service.get_by_id(Proveedor, db, factura.contraparte_id)
        proveedor_id = contraparte.id
        # enviar del fron indicador de pdv

    if tipo_contraparte.descripcion == TipoContraparteEnum.PROPIETARIO.value:
        contraparte = service.get_by_id(Propietario, db, factura.contraparte_id)
        propietario_id = contraparte.id

    if tipo_contraparte.descripcion == TipoContraparteEnum.REMITENTE.value:
        contraparte = service.get_by_id(Remitente, db, factura.contraparte_id)
        remitente_id = contraparte.id

    # TODO: ver caso mov factura contraparte otros
    # TODO: ver caso PDV

    create_movimiento(
        db,
        MovimientoForm(
            liquidacion_id=factura.liquidacion_id,
            tipo_contraparte_id=tipo_contraparte.id,
            contraparte_id=contraparte.id,

            chofer_id=chofer_id,
            proveedor_id=proveedor_id,
            propietario_id=propietario_id,
            remitente_id=remitente_id,
            punto_venta_id=punto_venta_id,

            contraparte=factura.contribuyente,
            contraparte_numero_documento=factura.ruc,
            tipo_documento_relacionado_id=tipo_documento_relacionado.id,
            #numero_documento_relacionado=anticipo.orden_carga_id,
            cuenta_id=tipo_cuenta.id,
            tipo_movimiento_id=tipo_movimiento.id,
            estado=MovimientoEstadoEnum.EN_PROCESO,
            detalle=tipo_movimiento.descripcion,
            monto = factura.iva *-1 if factura.es_cobro else  factura.iva,
            moneda_id=factura.moneda_id,
            tipo_cambio_moneda=1,  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en Descuento  # noqa
            fecha= datetime.now(),
            fecha_cambio_moneda=datetime.now(),
            tipo_movimiento_info='IVA'
        ),
        gestor_carga_id,
        modified_by,
    )
    return create_movimiento(
        db,
        MovimientoForm(
            liquidacion_id=factura.liquidacion_id,
            tipo_contraparte_id=tipo_contraparte.id,
            contraparte_id=contraparte.id,
            propietario_id=contraparte.id,
            contraparte=factura.contribuyente,
            contraparte_numero_documento=factura.ruc,
            tipo_documento_relacionado_id=tipo_documento_relacionado.id,
            #numero_documento_relacionado=anticipo.orden_carga_id,
            cuenta_id=tipo_cuenta.id,
            tipo_movimiento_id=tipo_movimiento.id,
            estado=MovimientoEstadoEnum.EN_PROCESO,
            detalle=tipo_movimiento.descripcion,
            monto= factura.retencion *-1 if factura.es_cobro else factura.retencion,
            moneda_id=factura.moneda_id,
            tipo_cambio_moneda=1,  # TODO: poner el tipo de cambio correcto en cuando se maneje tipo de cambio en Descuento  # noqa
            fecha= datetime.now(),
            fecha_cambio_moneda=datetime.now(),
            tipo_movimiento_info='RETENCION'
        ),
        gestor_carga_id,
        modified_by,
    )


def delete_movimiento_by_factura(db: Session,
    liquidacion: Liquidacion,
    modified_by: str,
    gestor_carga_id: Optional[int] = None,
) :

    tipo_movimiento = repositories.get_tipo_movimiento_by_descripcion(
        db, TipoMovimientoEnum.FISCAL.value
    )

    for c in liquidacion.movimientos:
        logger.info(f"delete mov => {c.id} - {c.tipo_movimiento_descripcion} - {c.tipo_movimiento_info}")
        if ( c.tipo_movimiento_descripcion == tipo_movimiento.descripcion
                and ( c.tipo_movimiento_info == 'IVA' or c.tipo_movimiento_info == 'RETENCION')):
            delete_movimiento(db, c.id, modified_by)

