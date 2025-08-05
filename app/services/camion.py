import os
from typing import List, Optional

from app.models.combinacion import Combinacion
from fastapi import HTTPException, UploadFile, status  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories, schemas
from app.config import REPORTS_FOLDER
from app.enums import EstadoEnum
from app.models import Camion
from app.utils import get_gestor_carga_by_params

from .camion_check_files import check_files


def check_combinaciones_activas(db: Session) -> List[Camion]:
    camiones = repositories.get_camion_list(db)

    for camion in camiones:
        combinacion_tracto = db.query(Combinacion).filter(Combinacion.camion_id == camion.id).first()

        if combinacion_tracto and combinacion_tracto.estado != EstadoEnum.INACTIVO.value:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"La combinación de tracto para el camión {camion.placa} ya está activa."
            )

    return camiones


async def create_camion(
    db: Session,
    data: schemas.CamionForm,
    foto_file: Optional[UploadFile],
    foto_habilitacion_municipal_frente_file: Optional[UploadFile],
    foto_habilitacion_municipal_reverso_file: Optional[UploadFile],
    foto_habilitacion_transporte_frente_file: Optional[UploadFile],
    foto_habilitacion_transporte_reverso_file: Optional[UploadFile],
    foto_habilitacion_automotor_frente_file: Optional[UploadFile],
    foto_habilitacion_automotor_reverso_file: Optional[UploadFile],
    gestor_carga_id: Optional[int],
    modified_by: str,
    usuario_id: int,
) -> schemas.Camion:
    gestor_id = get_gestor_carga_by_params(data, gestor_carga_id)
    gestor = repositories.get_gestor_carga_by_id(db, gestor_id)
    if not gestor:
        raise HTTPException(status_code=409, detail="Debe elegir un Gestor de carga")
    if repositories.get_camion_by(db, data.placa):
        raise HTTPException(
            status_code=409,
            detail=f"El Tracto con placa {data.placa} ya existe",
        )
    (
        foto_url,
        foto_habilitacion_municipal_frente_url,
        foto_habilitacion_municipal_reverso_url,
        foto_habilitacion_transporte_frente_url,
        foto_habilitacion_transporte_reverso_url,
        foto_habilitacion_automotor_frente_url,
        foto_habilitacion_automotor_reverso_url,
    ) = await check_files(
        foto_file,
        foto_habilitacion_municipal_frente_file,
        foto_habilitacion_municipal_reverso_file,
        foto_habilitacion_transporte_frente_file,
        foto_habilitacion_transporte_reverso_file,
        foto_habilitacion_automotor_frente_file,
        foto_habilitacion_automotor_reverso_file,
    )
    if data.limite_cantidad_oc_activas is None:
        data.limite_cantidad_oc_activas = gestor.limite_cantidad_oc_activas
    return repositories.create_camion(
        db,
        data,
        foto_url,
        foto_habilitacion_municipal_frente_url,
        foto_habilitacion_municipal_reverso_url,
        foto_habilitacion_transporte_frente_url,
        foto_habilitacion_transporte_reverso_url,
        foto_habilitacion_automotor_frente_url,
        foto_habilitacion_automotor_reverso_url,
        modified_by,
        gestor_carga_id,
        usuario_id,
    )


def get_camion_by_id(db: Session, id: int) -> Camion:
    obj = repositories.get_camion_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Tracto no encontrado")
    return obj


async def edit_camion(
    id: int,
    db: Session,
    data: schemas.CamionForm,
    foto_file: Optional[UploadFile],
    foto_habilitacion_municipal_frente_file: Optional[UploadFile],
    foto_habilitacion_municipal_reverso_file: Optional[UploadFile],
    foto_habilitacion_transporte_frente_file: Optional[UploadFile],
    foto_habilitacion_transporte_reverso_file: Optional[UploadFile],
    foto_habilitacion_automotor_frente_file: Optional[UploadFile],
    foto_habilitacion_automotor_reverso_file: Optional[UploadFile],
    modified_by: str,
) -> schemas.Camion:
    if data.placa:
        exists = repositories.get_camion_by(db, data.placa)
        if exists and exists.id != id:
            raise HTTPException(
                status_code=409,
                detail=f"El Tracto con placa {data.placa} ya existe",
            )
    (
        foto_url,
        foto_habilitacion_municipal_frente_url,
        foto_habilitacion_municipal_reverso_url,
        foto_habilitacion_transporte_frente_url,
        foto_habilitacion_transporte_reverso_url,
        foto_habilitacion_automotor_frente_url,
        foto_habilitacion_automotor_reverso_url,
    ) = await check_files(
        foto_file,
        foto_habilitacion_municipal_frente_file,
        foto_habilitacion_municipal_reverso_file,
        foto_habilitacion_transporte_frente_file,
        foto_habilitacion_transporte_reverso_file,
        foto_habilitacion_automotor_frente_file,
        foto_habilitacion_automotor_reverso_file,
    )
    to_edit_obj = get_camion_by_id(db, id)
    return repositories.edit_camion(
        to_edit_obj,
        db,
        data,
        foto_url,
        foto_habilitacion_municipal_frente_url,
        foto_habilitacion_municipal_reverso_url,
        foto_habilitacion_transporte_frente_url,
        foto_habilitacion_transporte_reverso_url,
        foto_habilitacion_automotor_frente_url,
        foto_habilitacion_automotor_reverso_url,
        modified_by,
    )


def delete_camion(db: Session, id: int, modified_by: str) -> schemas.Camion:
    co = get_camion_by_id(db, id)
    return repositories.delete_camion(co, db, modified_by)


def change_camion_status(
    db: Session, id: int, status: EstadoEnum, modified_by: str
) -> schemas.Camion:
    camion = get_camion_by_id(db, id)
    if not camion:
        raise HTTPException(status_code=404, detail="Camión no encontrado.")

    repositories.change_camion_status(camion, db, status, modified_by)

    combinaciones_relacionadas = repositories.get_combinaciones_by_camion_id(db, camion.id)

    if status == EstadoEnum.INACTIVO:
        # Inactivar cada combinación asociada
        for combinacion in combinaciones_relacionadas:
            repositories.change_combinacion_status(combinacion, db, EstadoEnum.INACTIVO, modified_by)

    elif status == EstadoEnum.ACTIVO:
        for combinacion in combinaciones_relacionadas:
            if (
                combinacion.semi and combinacion.semi.estado == EstadoEnum.ACTIVO.value and
                combinacion.chofer and combinacion.chofer.estado == EstadoEnum.ACTIVO.value and
                combinacion.propietario and combinacion.propietario.estado == EstadoEnum.ACTIVO.value
            ):
                # Verificar que no exista otra combinación activa con el mismo tracto y propietario
                conflicto = db.query(Combinacion).filter(
                    Combinacion.id != combinacion.id,
                    Combinacion.camion_id == combinacion.camion_id,
                    Combinacion.propietario_id == combinacion.propietario_id,
                    Combinacion.estado != EstadoEnum.INACTIVO.value
                ).first()

                if not conflicto:
                    repositories.change_combinacion_status(combinacion, db, EstadoEnum.ACTIVO, modified_by)

    db.commit()
    return schemas.Camion.from_orm(camion)


def update_camion_anticipo_retirado(db: Session, camion: Camion):
    # Se actualiza el total_anticipos_retirados en camion
    camion.total_anticipos_retirados_en_estado_pendiente_o_en_proceso = (
        repositories.get_total_anticipo_retirado_by_camion_id(db, camion.id)
    )
    db.commit()


def get_camion_reports(db: Session) -> str:
    datalist = repositories.get_camion_list(db)
    wb = Workbook()
    # get worksheet
    ws = wb.active

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Placa"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Nombre del Propietario"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "RUC del Propietario"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Nombre del Chofer"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Número de Documento del Chofer"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=6)
    title_cell.value = "Número de Chasís"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Tipo"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=8)
    title_cell.value = "Marca"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=9)
    title_cell.value = "Gestor de Cuenta"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=10)
    title_cell.value = "Oficial de Cuenta"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=11)
    title_cell.value = "Estado"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=12)
    title_cell.value = "Usuario creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=13)
    title_cell.value = "Fecha creación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=14)
    title_cell.value = "Usuario modificación"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=15)
    title_cell.value = "Fecha modificación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.placa

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.propietario.nombre

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.propietario.ruc

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.chofer_nombre

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = item.chofer_numero_documento

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.numero_chasis

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = item.tipo_descripcion

        value_cell = ws.cell(row=row + 2, column=8)
        value_cell.value = item.marca_descripcion

        value_cell = ws.cell(row=row + 2, column=9)
        value_cell.value = item.gestor_cuenta_nombre

        value_cell = ws.cell(row=row + 2, column=10)
        value_cell.value = item.oficial_cuenta_nombre

        value_cell = ws.cell(row=row + 2, column=11)
        value_cell.value = item.estado

        value_cell = ws.cell(row=row + 2, column=12)
        value_cell.value = item.created_by

        value_cell = ws.cell(row=row + 2, column=13)
        value_cell.value = item.created_at

        value_cell = ws.cell(row=row + 2, column=14)
        value_cell.value = item.modified_by

        value_cell = ws.cell(row=row + 2, column=15)
        value_cell.value = item.modified_at

    ws.auto_filter.ref = ws.dimensions
    filename = "camion_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
