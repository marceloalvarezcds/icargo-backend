import os
from typing import List, Optional

from fastapi import HTTPException, UploadFile  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore
from sqlalchemy.orm import Session  # type: ignore

from app import repositories, schemas
from app.config import REPORTS_FOLDER
from app.models import (
    GestorCargaPuntoVenta,
    InsumoPuntoVenta,
    PuntoVenta,
    PuntoVentaContactoGestorCarga,
)

from .gestor_carga_punto_venta import (
    create_gestor_carga_punto_venta,
    edit_gestor_carga_punto_venta,
)
from .pictshare import upload_and_get_image_url
from .punto_venta_contacto import update_punto_venta_contacto_list


def get_punto_venta_detail(
    db: Session, obj: PuntoVenta, gestor_carga_id: Optional[int]
) -> schemas.PuntoVenta:
    schema = schemas.PuntoVenta.from_orm(obj)
    contactos: List[
        PuntoVentaContactoGestorCarga
    ] = repositories.get_punto_venta_contacto_gestor_carga_list_by_punto_venta_id(
        db, obj.id, gestor_carga_id
    )
    ges: List[GestorCargaPuntoVenta] = obj.gestores
    gestores = [x for x in ges if x.gestor_carga_id == gestor_carga_id]
    insumos: List[InsumoPuntoVenta] = obj.insumos
    schema.contactos = [
        schemas.PuntoVentaContactoGestorCargaList.from_orm(x)
        for x in contactos
        if x.gestor_carga_id == gestor_carga_id
    ]
    schema.gestor_carga_punto_venta = gestores[0] if len(gestores) > 0 else None
    schema.insumos = [x for x in insumos if x.gestor_carga_id == gestor_carga_id]
    return schema


async def create_punto_venta(
    db: Session,
    data: schemas.PuntoVentaForm,
    file: Optional[UploadFile],
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> schemas.PuntoVenta:
    punto_existente = db.query(PuntoVenta).filter(
        PuntoVenta.numero_documento == data.numero_documento,
        PuntoVenta.proveedor_id == data.proveedor_id,
        PuntoVenta.numero_sucursal == data.numero_sucursal
    ).first()

    if punto_existente:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Ya existe un Punto de Venta con documento {data.numero_documento} "
                f"y sucursal número {data.numero_sucursal} para el proveedor especificado"
            ),
        )

    logo_url = await upload_and_get_image_url(file)
    obj = repositories.create_punto_venta(db, data, logo_url, modified_by)
    update_punto_venta_contacto_list(
        db, data.contactos, obj, gestor_carga_id, modified_by
    )
    create_gestor_carga_punto_venta(db, obj, gestor_carga_id, data.alias, modified_by)
    return get_punto_venta_detail(db, obj, gestor_carga_id)


def get_punto_venta_by_id(db: Session, id: int) -> PuntoVenta:
    obj = repositories.get_punto_venta_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Punto de Venta no encontrado")
    return obj


def get_punto_venta_by_id_and_gestor_carga_id(
    db: Session, id: int, gestor_carga_id: Optional[int] = None
) -> schemas.PuntoVenta:
    obj = repositories.get_punto_venta_by_id(db, id)
    if not obj:
        raise HTTPException(status_code=404, detail="Punto de Venta no encontrado")
    return get_punto_venta_detail(db, obj, gestor_carga_id)


async def edit_punto_venta(
    id: int,
    db: Session,
    data: schemas.PuntoVentaForm,
    file: Optional[UploadFile],
    gestor_carga_id: Optional[int],
    modified_by: str,
) -> schemas.PuntoVenta:
    exists = repositories.get_punto_venta_by(
        db, data.tipo_documento_id, data.numero_documento
    )
    if exists and exists.id != id:
        raise HTTPException(
            status_code=409,
            detail=f"El Punto de Venta con documento {data.numero_documento} ya existe",
        )

    existsSucursal = repositories.get_punto_venta_by_proveedor_sucursal(
        db, data.proveedor_id, data.numero_sucursal
    )
    if existsSucursal and existsSucursal.id != id:
        raise HTTPException(
            status_code=409,
            detail=f"El Punto de Venta numero {data.numero_sucursal} ya existe",
        )


    logo_url = await upload_and_get_image_url(file) if file else None
    to_edit_obj = get_punto_venta_by_id(db, id)
    obj = repositories.edit_punto_venta(to_edit_obj, db, data, logo_url, modified_by)
    update_punto_venta_contacto_list(
        db, data.contactos, obj, gestor_carga_id, modified_by
    )
    edit_gestor_carga_punto_venta(db, obj, gestor_carga_id, data.alias, modified_by)
    return get_punto_venta_detail(db, obj, gestor_carga_id)


def delete_punto_venta(
    db: Session, id: int, gestor_carga_id: Optional[int], modified_by: str
) -> schemas.PuntoVenta:
    co = get_punto_venta_by_id(db, id)
    obj = repositories.delete_punto_venta(co, db, modified_by)
    return get_punto_venta_detail(db, obj, gestor_carga_id)


def get_punto_venta_reports(db: Session, proveedor_id: int) -> str:
    datalist = repositories.get_punto_venta_list(db, proveedor_id)
    wb = Workbook()
    # get worksheet
    ws = wb.active

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Nombre"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Nombre de Fantasía"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "Tipo de Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=4)
    title_cell.value = "Número de Documento"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Composición Jurídica"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=6)
    title_cell.value = "Dirección"
    title_cell.font = Font(bold=True)

    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "Ubicación"
    title_cell.font = Font(bold=True)

    for row, item in enumerate(datalist):
        value_cell = ws.cell(row=row + 2, column=1)
        value_cell.value = item.nombre

        value_cell = ws.cell(row=row + 2, column=2)
        value_cell.value = item.nombre_corto if item.nombre_corto else ""

        value_cell = ws.cell(row=row + 2, column=3)
        value_cell.value = item.tipo_documento.descripcion

        value_cell = ws.cell(row=row + 2, column=4)
        value_cell.value = item.numero_documento

        value_cell = ws.cell(row=row + 2, column=5)
        value_cell.value = (
            item.composicion_juridica.nombre if item.composicion_juridica else ""
        )

        value_cell = ws.cell(row=row + 2, column=6)
        value_cell.value = item.direccion if item.direccion else ""

        value_cell = ws.cell(row=row + 2, column=7)
        value_cell.value = (
            f"{item.ciudad.nombre}/{item.ciudad.localidad.nombre}/{item.ciudad.localidad.pais.nombre_corto}"  # noqa: B950
            if item.ciudad
            else ""
        )

    ws.auto_filter.ref = ws.dimensions
    filename = "punto_venta_reports.xls"
    # Save the file
    wb.save(os.path.join(REPORTS_FOLDER, filename))
    return filename
